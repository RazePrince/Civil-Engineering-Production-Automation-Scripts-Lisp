"""
BBS Transformer - Converts Armabeton XLS (Nomenclature) to
English Bar Bending Schedule XLSX.

Usage:
  python bbs_transformer.py                   <- GUI
  python bbs_transformer.py src.xls out.xlsx  <- CLI

Sketch options:
  - Default: preserve sketch aspect ratio and center it in the SKETCH cell.
  - Exact fit: stretch the sketch to the full SKETCH cell when requested.

This version keeps the previous workbook format but makes the sketch-placement
options explicit and easier to follow in the CLI, batch flow, and GUI.
"""

import re, sys, os, threading
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from io import BytesIO

# ── styles ──────────────────────────────────────────────────────────────────
T = Side(style="thin")
ALL_T = Border(left=T, right=T, top=T, bottom=T)

FONT_C11  = dict(name="Calibri", size=11)
FONT_A9B  = dict(name="Arial",   size=9,  bold=True)
FONT_A10R = dict(name="Arial",   size=10, bold=True, color="FFFF0000")
FONT_A10  = dict(name="Arial",   size=10)

FILL_GREY = PatternFill("solid", fgColor="FFC0C0C0")
FILL_DARK = PatternFill("solid", fgColor="FF808080")

CTR  = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left",   vertical="center")

# ── Layout (NO gap column):
#   A          = label           (col 1)
#   B:L merged = value           (cols 2-12)
#   M:N merged = right label     (cols 13-14)
#   O          = right value     (col 15)
#   P+         = dark grey (hidden area)
#
# Data columns shift to use A-O for BBS table (same 15 cols)

COL_WIDTHS = {
    "A": 20,    # label
    "B": 50,    # value (merged B:L — wide enough for long names)
    "C": 0.5, "D": 0.5, "E": 0.5, "F": 0.5,
    "G": 0.5, "H": 0.5, "I": 0.5, "J": 0.5,
    "K": 0.5, "L": 0.5,
    "M": 22,    # right label
    "N": 0.5,   # right label overflow (merged with M)
    "O": 14,    # right value
}

# BBS data table columns (target Armabeton-style English layout)
# A:Q are the visible BBS table, P/R/S are kept hidden to match the source workbook structure.
BBS_COL_WIDTHS = {
    "A": 19.86, "B": 11.00, "C": 9.71,  "D": 12.57, "E": 9.71,
    "F": 12.57, "G": 13.00, "H": 15.29, "I": 16.71, "J": 10.71,
    "K": 13.00, "L": 13.00, "M": 13.00, "N": 13.00, "O": 29.43,
    "P": 13.43, "Q": 42.00, "R": 20.29, "S": 8.14,
}

BBS_COL_HDRS = [
    "MARK", "DIAMETER", "NO. IN EACH", "NO. OF ELEMENTS", "TOTAL NO. OF BARS",
    "SPACING ", "SHAPE CODE", '"A"', '"B"', '"C"', '"D"', '"E"', '"F"', '"G"',
    "TOTAL LENGTH", "TOTAL LENGTH", "SKETCH", "ARMACAD LENGTH",
]

SUPPORT_SHEETS_TO_COPY = []  # No support sheets are copied into the output


# ── sketch ──────────────────────────────────────────────────────────────────
def _numval(v):
    try: return float(str(v).split("->")[0].strip())
    except: return 0

def draw_sketch(dims, bar_type=None, total_len=None):
    """
    Draw a schematic bar sketch for GUI preview and for generated Excel sketches.

    v10 note:
    - Better orientation for stirrups/U-bars and crank bars.
    - Rounded bends instead of sharp staircase-only shapes.
    - Schematic scaling for bars with very uneven dimensions, so hooks stay visible.
    - Keeps the old call signature: draw_sketch(dims) still works.
    - Detects closed stirrups such as bar 451 (A/B/C with C as hook length).
    """
    try:
        from PIL import Image as PI, ImageDraw as PD, ImageFont as PF
    except ImportError:
        return None

    W, H = 600, 220
    SCALE_AA = 3  # antialiasing for smoother fillet-like bends
    img = PI.new("RGBA", (W*SCALE_AA, H*SCALE_AA), (255, 255, 255, 0))
    draw = PD.Draw(img)
    RED = (200, 0, 0, 255)
    GREY = (75, 75, 75, 190)
    lw = 5 * SCALE_AA
    PAD_L = 55 * SCALE_AA; PAD_R = 35 * SCALE_AA
    PAD_T = 34 * SCALE_AA; PAD_B = 32 * SCALE_AA

    try:
        font = PF.truetype("arial.ttf", 18*SCALE_AA)
    except Exception:
        try:
            font = PF.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 18*SCALE_AA)
        except Exception:
            font = PF.load_default()

    def tw(text):
        bb = draw.textbbox((0, 0), text, font=font)
        return bb[2]-bb[0], bb[3]-bb[1]

    def num(k):
        return _numval(dims.get(k, 0))

    def label_text(k):
        return f"{k}={dims.get(k, '')}" if dims.get(k, '') not in (None, "") else k

    keys = [k for k in ["A", "B", "C", "D", "E"] if k in dims]
    n = len(keys)
    nv = {k: max(1.0, num(k)) for k in keys}

    # Helpers -----------------------------------------------------------------
    def schematic_lengths(vals, min_len=0.28, max_len=2.6):
        """Compress extreme lengths so small hooks stay visible in preview."""
        vals = [max(1.0, float(v)) for v in vals]
        m = max(vals) if vals else 1.0
        # sqrt compression keeps proportions but avoids 8000-mm segment hiding 230-mm hooks
        return [min(max_len, max(min_len, (v/m)**0.5 * max_len)) for v in vals]

    def fit_points(points):
        xs = [p[0] for p in points]; ys = [p[1] for p in points]
        minx, maxx = min(xs), max(xs); miny, maxy = min(ys), max(ys)
        ww = max(0.001, maxx-minx); hh = max(0.001, maxy-miny)
        avail_w = W*SCALE_AA - PAD_L - PAD_R
        avail_h = H*SCALE_AA - PAD_T - PAD_B
        s = min(avail_w/ww, avail_h/hh)
        # avoid giant zoom on tiny/schematic shapes
        s = min(s, 130*SCALE_AA)
        ox = PAD_L + (avail_w - ww*s)/2 - minx*s
        oy = PAD_T + (avail_h - hh*s)/2 - miny*s
        return [(int(ox + x*s), int(oy + y*s)) for x, y in points]

    def rounded_polyline(points, fill=RED, width=lw):
        if len(points) < 2:
            return
        try:
            draw.line(points, fill=fill, width=width, joint="curve")
        except TypeError:
            draw.line(points, fill=fill, width=width)
        # round caps/joints
        r = width // 2
        for x, y in points:
            draw.ellipse((x-r, y-r, x+r, y+r), fill=fill)

    def add_segment_label(k, p1, p2, side=1):
        x1, y1 = p1; x2, y2 = p2
        mx, my = (x1+x2)//2, (y1+y2)//2
        txt = label_text(k)
        w, h = tw(txt)
        dx, dy = x2-x1, y2-y1
        # normal offset
        if abs(dx) >= abs(dy):
            offx, offy = 0, -18*SCALE_AA if side >= 0 else 18*SCALE_AA
            ax = mx - w//2; ay = my + offy - h//2
        else:
            offx, offy = -26*SCALE_AA if side >= 0 else 26*SCALE_AA, 0
            ax = mx + offx - w//2; ay = my + offy - h//2
        ax = max(2*SCALE_AA, min(ax, W*SCALE_AA-w-2*SCALE_AA))
        ay = max(2*SCALE_AA, min(ay, H*SCALE_AA-h-2*SCALE_AA))
        draw.text((ax, ay), txt, fill=RED, font=font)

    def render(path, label_segments=None, label_sides=None):
        """Fit a schematic polyline to the preview canvas and label its segments.

        label_segments is a list like [("A", 0), ("B", 1)], where the
        second value is the index of the segment between fitted points. This
        small helper was added so every generated shape uses the same scaling,
        rounded bend rendering, and label placement.
        """
        fitted = fit_points(path)
        rounded_polyline(fitted)
        label_segments = label_segments or []
        label_sides = label_sides or {}
        for key, seg_idx in label_segments:
            if 0 <= seg_idx < len(fitted) - 1:
                add_segment_label(key, fitted[seg_idx], fitted[seg_idx+1], label_sides.get(key, 1))

    def looks_like_closed_stirrup(A, B, Cc):
        """Closed stirrups in Armabeton often come as A=width, B=height, C=hook.
        Example bar 451: A=225, B=210, C=160, total length around 1106.
        The total length is much bigger than A+B+C because the closed link uses
        both sides of the rectangle plus hooks/bend allowances.
        """
        try:
            tl = float(total_len) if total_len not in (None, "", 0) else 0.0
        except Exception:
            tl = 0.0
        dim_sum = A + B + Cc
        # Closed links/stirrups usually have A and B as the box sides and C as
        # a hook/overlap. The reported total length is therefore closer to the
        # full perimeter plus hook allowance than to A+B+C.
        hook_reasonable = Cc <= max(A, B) * 0.95 and Cc >= min(A, B) * 0.35
        compact_link = max(A, B) <= 900 and min(A, B) <= 650
        perimeter_hint = (2*A + 2*B) * 0.85 <= tl <= (2*A + 2*B + 2*Cc) * 1.25 if tl else False
        strong_total_hint = tl > dim_sum * 1.45 if tl else False
        near_square_link = 0.45 <= A / max(B, 1.0) <= 2.25
        return hook_reasonable and compact_link and near_square_link and (perimeter_hint or strong_total_hint)

    def render_closed_stirrup():
        # A = horizontal clear width, B = vertical clear height, C = diagonal hook.
        A = nv.get("A", 1.0); B = nv.get("B", 1.0)
        ratio = max(0.55, min(2.2, (A / max(B, 1.0))))
        rw = min(3.15, max(2.25, 2.65 * ratio**0.35))
        rh = min(1.65, max(1.05, rw / max(ratio, 0.70)))
        pts = fit_points([(0, 0), (rw, rh)])
        x1, y1 = pts[0]; x2, y2 = pts[1]
        if y2 < y1:
            y1, y2 = y2, y1
        radius = max(12*SCALE_AA, min((x2-x1)//6, (y2-y1)//3))
        try:
            draw.rounded_rectangle((x1, y1, x2, y2), radius=radius, outline=RED, width=lw)
        except Exception:
            rounded_polyline([(x1+radius,y1),(x2-radius,y1),(x2,y1+radius),(x2,y2-radius),(x2-radius,y2),(x1+radius,y2),(x1,y2-radius),(x1,y1+radius),(x1+radius,y1)])

        # Two small diagonal stirrup hooks near upper-left, kept inside the link.
        hlen = min((x2-x1)*0.22, (y2-y1)*0.46)
        p1 = (int(x1 + (x2-x1)*0.10), int(y1 + (y2-y1)*0.05))
        p2 = (int(p1[0] + hlen*0.58), int(p1[1] + hlen*0.95))
        p3 = (int(x1 + (x2-x1)*0.04), int(y1 + (y2-y1)*0.26))
        p4 = (int(p3[0] + hlen*0.56), int(p3[1] + hlen*0.58))
        rounded_polyline([p1, p2])
        rounded_polyline([p3, p4])

        # Labels: A in the body, B to the right, C by the hook.
        def draw_label(txt, x, y):
            w, h = tw(txt)
            draw.text((max(2*SCALE_AA, min(int(x-w/2), W*SCALE_AA-w-2*SCALE_AA)),
                       max(2*SCALE_AA, min(int(y-h/2), H*SCALE_AA-h-2*SCALE_AA))), txt, fill=RED, font=font)
        draw_label(label_text("A"), (x1+x2)//2, (y1+y2)//2 + 8*SCALE_AA)
        draw_label(label_text("B"), x2 + 22*SCALE_AA, (y1+y2)//2)
        draw_label(label_text("C"), x1 + (x2-x1)*0.22, y1 + (y2-y1)*0.18)

    # Shape builders -----------------------------------------------------------
    if n == 0:
        y = H*SCALE_AA // 2
        rounded_polyline([(PAD_L, y), (W*SCALE_AA-PAD_R, y)])

    elif n == 1:
        render([(0, 0), (3.2, 0)], [("A", 0)])

    elif n == 2:
        A, B = nv["A"], nv["B"]
        la, lb = schematic_lengths([A, B], max_len=3.0)
        # Common hooked longitudinal bar: short A hook + long B run.
        if A <= B * 0.55:
            path = [(0, -la), (0, 0), (lb, 0)]
            render(path, [("A", 0), ("B", 1)], {"A": 1, "B": -1})
        elif B <= A * 0.55:
            path = [(0, 0), (la, 0), (la, -lb)]
            render(path, [("A", 0), ("B", 1)], {"A": 1, "B": -1})
        else:
            # balanced L bar
            path = [(0, -la), (0, 0), (lb, 0)]
            render(path, [("A", 0), ("B", 1)], {"A": 1, "B": -1})

    elif n == 3:
        A, B, Cc = nv["A"], nv["B"], nv["C"]
        la, lb, lc = schematic_lengths([A, B, Cc], max_len=2.8)

        # Closed stirrup/link: A=width, B=height, C=hook length.
        # This fixes cases like bar 451, which should not preview as a step/crank bar.
        if looks_like_closed_stirrup(A, B, Cc):
            render_closed_stirrup()

        # U / stirrup style: two similar hooks/legs around a middle run.
        elif abs(A-Cc) / max(A, Cc) < 0.35 and B >= min(A, Cc) * 0.9:
            path = [(0, -la), (0, 0), (lb, 0), (lb, -lc)]
            render(path, [("A", 0), ("B", 1), ("C", 2)], {"A": 1, "B": -1, "C": -1})

        # Small top hook + vertical drop + horizontal leg.
        elif A <= B * 0.55 and Cc <= B * 1.4:
            path = [(la, -lb), (0, -lb), (0, 0), (lc, 0)]
            render(path, [("A", 0), ("B", 1), ("C", 2)], {"A": 1, "B": 1, "C": -1})

        # Long left/right legs with a small crank between them.
        elif B <= max(A, Cc) * 0.35:
            path = [(0, 0), (la, 0), (la, lb), (la+lc, lb)]
            render(path, [("A", 0), ("B", 1), ("C", 2)], {"A": 1, "B": -1, "C": -1})

        # Long tail with small hook at far end.
        elif Cc >= max(A, B) * 1.8:
            path = [(0, -la), (0, 0), (lb, 0), (lb, 0.42), (lb+lc, 0.42)]
            render(path, [("A", 0), ("B", 1), ("C", 3)], {"A": 1, "B": -1, "C": -1})

        else:
            # Generic crank/Z bar.
            path = [(0, 0), (la, 0), (la, lb), (la+lc, lb)]
            render(path, [("A", 0), ("B", 1), ("C", 2)], {"A": 1, "B": -1, "C": -1})

    elif n == 4:
        A, B, Cc, D = nv["A"], nv["B"], nv["C"], nv["D"]
        la, lb, lc, ld = schematic_lengths([A, B, Cc, D], max_len=2.7)
        c_is_variable = "->" in str(dims.get("C", ""))

        if c_is_variable or Cc >= max(A, B, D) * 1.3:
            # Cranked longitudinal bar: A hook, B drop, variable/long C run, D end hook.
            path = [(0, 0), (la, 0), (la, lb), (la+lc, lb), (la+lc, lb-ld)]
            render(path, [("A", 0), ("B", 1), ("C", 2), ("D", 3)],
                   {"A": 1, "B": -1, "C": -1, "D": -1})
        else:
            # Open stirrup / U bar: A top hook, B left leg, C bottom run, D right leg.
            # Schematic proportions prevent huge beam lengths from destroying the stirrup shape.
            path = [(-0.42, -1.05), (0, -1.05), (0, 0), (2.55, 0), (2.55, -0.95)]
            render(path, [("A", 0), ("B", 1), ("C", 2), ("D", 3)],
                   {"A": 1, "B": 1, "C": -1, "D": -1})

    else:
        # 5 dimensions: typical stirrup / open box with two top hooks.
        path = [(-0.42, -1.05), (0, -1.05), (0, 0), (2.55, 0), (2.55, -0.95), (3.00, -0.95)]
        render(path, [("A", 0), ("B", 1), ("C", 2), ("D", 3), ("E", 4)],
               {"A": 1, "B": 1, "C": -1, "D": -1, "E": 1})

    # Downsample for antialiased edges
    img = img.resize((W, H), PI.LANCZOS)
    buf = BytesIO(); img.save(buf, format="PNG"); buf.seek(0)
    return buf

def parse_nomenclature(src_path):
    df = pd.read_excel(src_path, sheet_name="Nomenclature",
                       engine="openpyxl", header=None)
    rows = []
    for _, r in df.iterrows():
        v0 = r[0]
        if pd.isna(v0) or not str(v0).strip().lstrip("-").isdigit():
            continue
        repere   = str(r[1]).strip() if pd.notna(r[1]) else ""
        nb_elem  = int(r[2]) if pd.notna(r[2]) else 0
        nb_steel = int(r[3]) if pd.notna(r[3]) else 1
        nb_total = int(r[4]) if pd.notna(r[4]) else 0
        mandrin  = r[5] if pd.notna(r[5]) else ""
        # Armabeton puts the shape code in this column; the bar diameter is encoded in Repères (T10, T16, etc.).
        shape_raw = str(r[6]).strip() if pd.notna(r[6]) else ""
        lon_raw   = str(r[7]).strip() if pd.notna(r[7]) else ""
        esp       = r[8] if pd.notna(r[8]) else ""
        m = re.match(r"([A-Za-z]*)(\d+)", repere)
        bar_dia = int(m.group(2)) if m else int(float(re.sub(r"[^\d.]", "", shape_raw) or 0))
        try:
            total_len = int(float(str(esp).replace(",", "."))) \
                if str(esp) not in ("nan", "VARIES", "") else 0
        except Exception:
            total_len = 0
        dims = {}
        # New target format exposes A through G, not just A through E.
        for k, v in re.findall(r"([A-G])\s*=\s*([\d.]+(?:->[\d.]+)?)", lon_raw):
            dims[k] = v
        try:
            shape_code = float(str(shape_raw).replace(",", ".")) if shape_raw not in ("", "nan") else None
            if shape_code is not None and shape_code == int(shape_code):
                shape_code = int(shape_code)
        except Exception:
            shape_code = shape_raw or None
        rows.append({"idx": int(v0), "bar_type": repere, "nb_elem": nb_elem,
                     "nb_steel": nb_steel, "nb_total": nb_total, "mandrin": mandrin,
                     "diameter": bar_dia, "shape_code": shape_code, "dims": dims,
                     "total_len": total_len, "src_row": int(_) - 1})
    return pd.DataFrame(rows)


def _load_workbook_any_ext(path, **kwargs):
    """Load XLSX content even when Armabeton saved it with a .xls extension."""
    from openpyxl import load_workbook as _lwb
    try:
        return _lwb(path, **kwargs)
    except Exception:
        with open(path, "rb") as f:
            return _lwb(BytesIO(f.read()), **kwargs)


def parse_unit_weights(src_path):
    """Return {diameter_mm: exact_unit_weight} from the source Poids d'acier/TONNAGE sheet when present."""
    weights = {}
    fallback = {6:0.222, 8:0.395, 10:0.620, 12:0.888, 16:1.580,
                20:2.470, 25:3.854, 28:4.834, 32:6.310, 40:9.865}
    try:
        wb = _load_workbook_any_ext(src_path, data_only=False)
        sheet_name = None
        lower = {s.lower(): s for s in wb.sheetnames}
        for key in ("poids d'acier", "tonnage"):
            if key in lower:
                sheet_name = lower[key]
                break
        if sheet_name:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                dia = None
                # Source French sheet stores diameter in A; converted English sheet stores it in B.
                for cell in row[:2]:
                    try:
                        if cell.value not in (None, "") and float(cell.value) > 0:
                            dia = int(float(cell.value)); break
                    except Exception:
                        pass
                if not dia:
                    continue
                # Unit weight is column F in both source and target layouts.
                try:
                    w = row[5].value
                    if w not in (None, ""):
                        weights[dia] = float(str(w).replace(",", "."))
                except Exception:
                    pass
    except Exception:
        pass
    # Keep source values exact; only backfill missing common diameters.
    for d, w in fallback.items():
        weights.setdefault(d, w)
    return weights


def parse_couplers(src_path):
    """Parse the Armabeton Coupleurs sheet when it exists.

    Returns a dict with diameter columns, threaded-only counts, coupler type
    rows, and a flag showing whether any source coupler data was present.
    The source workbook can be true .xlsx content saved with a .xls extension.
    """
    try:
        wb = _load_workbook_any_ext(src_path, data_only=False)
    except Exception:
        return None

    sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() == "coupleurs":
            sheet_name = name
            break
    if not sheet_name:
        return None

    ws = wb[sheet_name]

    # Source layout normally uses C:L for diameter/count columns, with totals in M.
    diameter_row = None
    for r in range(1, min(ws.max_row, 30) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 20) + 1)]
        if any(str(v).strip().lower().startswith("diam") for v in row_vals if v is not None):
            diameter_row = r + 1
            break
    if not diameter_row:
        # Fallback for Armabeton Coupleurs default: header row 5, diameters row 6.
        diameter_row = 6

    first_dia_col = 3
    last_dia_col = 12
    diameters = []
    for c in range(first_dia_col, min(last_dia_col, ws.max_column) + 1):
        diameters.append(ws.cell(diameter_row, c).value)

    threaded_row = diameter_row + 1
    threaded_counts = [ws.cell(threaded_row, c).value for c in range(first_dia_col, min(last_dia_col, ws.max_column) + 1)]

    type_rows = []
    for r in range(threaded_row + 1, min(ws.max_row, threaded_row + 12) + 1):
        label = ws.cell(r, 1).value
        code = ws.cell(r, 2).value
        counts = [ws.cell(r, c).value for c in range(first_dia_col, min(last_dia_col, ws.max_column) + 1)]
        if label is None and code is None and not any(v not in (None, "") for v in counts):
            continue
        if str(label or "").strip().lower() == "type" or code not in (None, ""):
            type_rows.append({"label": "Type", "code": code, "counts": counts})

    def numeric_present(values):
        for v in values:
            try:
                if v not in (None, "") and float(v) != 0:
                    return True
            except Exception:
                pass
        return False

    has_data = any(v not in (None, "") for v in diameters) or numeric_present(threaded_counts)
    has_data = has_data or any(numeric_present(row["counts"]) or row.get("code") not in (None, "") for row in type_rows)
    if not has_data:
        return None

    return {
        "diameters": diameters,
        "threaded_counts": threaded_counts,
        "type_rows": type_rows,
    }


def write_couplers_on_tonnage(ws, coupler_data, start_col=12, start_row=4):
    """Write a translated COUPLERS block to the right side of TONNAGE."""
    if not coupler_data:
        return

    from openpyxl.utils import get_column_letter

    diameters = list(coupler_data.get("diameters") or [])
    threaded_counts = list(coupler_data.get("threaded_counts") or [])
    type_rows = list(coupler_data.get("type_rows") or [])

    # Keep at least 10 source diameter slots so the shape matches Armabeton.
    n_dias = max(10, len(diameters), len(threaded_counts), *(len(r.get("counts", [])) for r in type_rows) if type_rows else [0])
    diameters += [None] * (n_dias - len(diameters))
    threaded_counts += [None] * (n_dias - len(threaded_counts))
    for row in type_rows:
        row["counts"] = list(row.get("counts") or []) + [None] * (n_dias - len(row.get("counts") or []))

    col_label = start_col
    col_code = start_col + 1
    col_first = start_col + 2
    col_last = col_first + n_dias - 1
    col_total = col_last + 1
    last_col_letter = get_column_letter(col_total)

    FILL_GREY = PatternFill("solid", fgColor="FFD9D9D9")
    FONT_TITLE = Font(name="Calibri", size=20, bold=True, color="FF0070C0")
    FONT_HDR = Font(name="Calibri", size=9, bold=True)
    FONT_NORM = Font(name="Arial", size=10)
    FONT_GREEN = Font(name="Arial", size=10, bold=True, color="FF00A050")
    FONT_RED = Font(name="Arial", size=10, bold=True, color="FFFF0000")
    ALIGN_CTR = Alignment(horizontal="center", vertical="center")
    ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
    THIN = Side(style="thin")
    MED = Side(style="medium")
    DOTTED = Side(style="dotted")
    BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BORD_MED = Border(left=MED, right=MED, top=MED, bottom=MED)

    def put(r, c, value=None, font=FONT_NORM, fill=None, align=ALIGN_CTR, border=BORD, fmt=None):
        cell = ws.cell(r, c, value)
        cell.font = font
        cell.alignment = align
        cell.border = border
        if fill:
            cell.fill = fill
        if fmt:
            cell.number_format = fmt
        return cell

    # Column widths for the translated block.
    ws.column_dimensions[get_column_letter(col_label)].width = 18
    ws.column_dimensions[get_column_letter(col_code)].width = 8
    for c in range(col_first, col_last + 1):
        ws.column_dimensions[get_column_letter(c)].width = 8
    ws.column_dimensions[get_column_letter(col_total)].width = 10

    # Title.
    ws.merge_cells(start_row=start_row, start_column=col_label, end_row=start_row, end_column=col_total)
    put(start_row, col_label, "COUPLERS", FONT_TITLE, FILL_GREY, ALIGN_CTR, BORD_MED)
    for c in range(col_label + 1, col_total + 1):
        ws.cell(start_row, c).fill = FILL_GREY
        ws.cell(start_row, c).border = BORD_MED

    # Diameter header and diameter values.
    dia_header_row = start_row + 3
    dia_value_row = start_row + 4
    threaded_row = start_row + 5
    ws.merge_cells(start_row=dia_header_row, start_column=col_first, end_row=dia_header_row, end_column=col_last)
    put(dia_header_row, col_first, "Diameters", FONT_HDR, None, ALIGN_CTR, BORD)
    for c in range(col_first + 1, col_last + 1):
        ws.cell(dia_header_row, c).border = BORD
    for i, d in enumerate(diameters):
        put(dia_value_row, col_first + i, d, FONT_HDR, None, ALIGN_CTR, BORD)
    put(dia_value_row, col_label, None, FONT_NORM, None, ALIGN_CTR, BORD)
    put(dia_value_row, col_code, None, FONT_NORM, None, ALIGN_CTR, BORD)
    put(dia_value_row, col_total, None, FONT_NORM, None, ALIGN_CTR, BORD)

    # Threaded-only row.
    put(threaded_row, col_label, "Threaded ends only:", FONT_NORM, None, ALIGN_LEFT, BORD)
    put(threaded_row, col_code, None, FONT_NORM, None, ALIGN_CTR, BORD)
    for i, value in enumerate(threaded_counts):
        put(threaded_row, col_first + i, value if value not in (None, "") else None, FONT_NORM, None, ALIGN_CTR, BORD)
    first_l = get_column_letter(col_first)
    last_l = get_column_letter(col_last)
    total_l = get_column_letter(col_total)
    put(threaded_row, col_total, f"=SUM({first_l}{threaded_row}:{last_l}{threaded_row})", FONT_RED, None, ALIGN_CTR, BORD)

    # Coupler type rows.
    data_rows = [threaded_row]
    row = threaded_row + 1
    if not type_rows:
        type_rows = [{"label": "Type", "code": "", "counts": [0] * n_dias}]
    for item in type_rows:
        put(row, col_label, "Type", FONT_NORM, None, ALIGN_LEFT, Border(left=THIN, right=THIN, top=DOTTED, bottom=DOTTED))
        put(row, col_code, item.get("code"), FONT_GREEN, None, ALIGN_CTR, Border(left=THIN, right=THIN, top=DOTTED, bottom=DOTTED))
        for i, value in enumerate(item.get("counts") or []):
            put(row, col_first + i, value if value not in (None, "") else None, FONT_NORM, None, ALIGN_CTR,
                Border(left=THIN, right=THIN, top=DOTTED, bottom=DOTTED))
        put(row, col_total, f"=SUM({first_l}{row}:{last_l}{row})", FONT_RED, None, ALIGN_CTR,
            Border(left=THIN, right=THIN, top=DOTTED, bottom=DOTTED))
        data_rows.append(row)
        row += 1

    # Bottom border under the table.
    for c in range(col_label, col_total + 1):
        old = ws.cell(row - 1, c).border
        ws.cell(row - 1, c).border = Border(left=old.left, right=old.right, top=old.top, bottom=THIN)

    # Total number box below table.
    box_row = row + 3
    box_first = col_first + 2
    box_last = min(col_first + 5, col_total - 2)
    ws.merge_cells(start_row=box_row, start_column=box_first, end_row=box_row, end_column=box_last)
    ws.merge_cells(start_row=box_row + 1, start_column=box_first, end_row=box_row + 1, end_column=box_last)
    put(box_row, box_first, "Total number", FONT_RED, None, ALIGN_CTR, BORD)
    put(box_row + 1, box_first, f"=SUM({total_l}{data_rows[0]}:{total_l}{data_rows[-1]})", FONT_RED, None, ALIGN_CTR, BORD)
    for c in range(box_first + 1, box_last + 1):
        ws.cell(box_row, c).border = BORD
        ws.cell(box_row + 1, c).border = BORD



def copy_support_sheets(src_path, dst_wb):
    """Copy the supporting Armabeton sheets into the output workbook, preserving values and cell formatting."""
    from copy import copy as _copy
    copied = []
    try:
        src_wb = _load_workbook_any_ext(src_path, data_only=False)
    except Exception as e:
        print(f"  [support sheet copy warning: {e}]")
        return copied

    for name in SUPPORT_SHEETS_TO_COPY:
        if name not in src_wb.sheetnames:
            continue
        src_ws = src_wb[name]
        if name in dst_wb.sheetnames:
            dst_ws = dst_wb[name]
        else:
            dst_ws = dst_wb.create_sheet(name)
        dst_ws.sheet_properties.tabColor = src_ws.sheet_properties.tabColor
        dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
        dst_ws.freeze_panes = src_ws.freeze_panes

        for col_letter, dim in src_ws.column_dimensions.items():
            dst_dim = dst_ws.column_dimensions[col_letter]
            dst_dim.width = dim.width
            dst_dim.hidden = dim.hidden
            dst_dim.outlineLevel = dim.outlineLevel
            dst_dim.collapsed = dim.collapsed
        for row_idx, dim in src_ws.row_dimensions.items():
            dst_dim = dst_ws.row_dimensions[row_idx]
            dst_dim.height = dim.height
            dst_dim.hidden = dim.hidden
            dst_dim.outlineLevel = dim.outlineLevel
            dst_dim.collapsed = dim.collapsed

        for merged in src_ws.merged_cells.ranges:
            dst_ws.merge_cells(str(merged))

        for row in src_ws.iter_rows():
            for src_cell in row:
                dst_cell = dst_ws.cell(row=src_cell.row, column=src_cell.column, value=src_cell.value)
                if src_cell.has_style:
                    dst_cell.font = _copy(src_cell.font)
                    dst_cell.fill = _copy(src_cell.fill)
                    dst_cell.border = _copy(src_cell.border)
                    dst_cell.alignment = _copy(src_cell.alignment)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.protection = _copy(src_cell.protection)
                if src_cell.comment:
                    dst_cell.comment = _copy(src_cell.comment)
                if src_cell.hyperlink:
                    dst_cell._hyperlink = _copy(src_cell.hyperlink)
        copied.append(name)

    # A few copied support sheets refer to nb_elm; define it as a constant so formulas recalculate cleanly.
    try:
        from openpyxl.workbook.defined_name import DefinedName
        if "nb_elm" not in dst_wb.defined_names:
            dst_wb.defined_names.add(DefinedName("nb_elm", attr_text="1"))
    except Exception:
        pass
    return copied




def _trim_transparent_pil_image(pil_img, pad_px=8):
    """Crop transparent padding from a generated sketch image, keeping a small margin."""
    try:
        if pil_img.mode != "RGBA":
            pil_img = pil_img.convert("RGBA")
        bbox = pil_img.getchannel("A").getbbox()
        if not bbox:
            return pil_img
        l, t, r, b = bbox
        l = max(0, l - int(pad_px))
        t = max(0, t - int(pad_px))
        r = min(pil_img.width, r + int(pad_px))
        b = min(pil_img.height, b + int(pad_px))
        return pil_img.crop((l, t, r, b))
    except Exception:
        return pil_img


def _trim_preview_sketch_buffer(buf, pad_px=10):
    """Return a PNG buffer with transparent preview-canvas padding removed.

    The GUI/Excel preview sketch generator draws into a 600x220 transparent
    canvas. If that full canvas is inserted into the BBS SKETCH cell, the actual
    bar can look small or off-center. Cropping transparent padding first makes
    the generated preview sketch fit like the source WMF sketches while still
    preserving aspect ratio.
    """
    try:
        from PIL import Image as _PILImage
        buf.seek(0)
        img = _PILImage.open(buf).convert("RGBA")
        img = _trim_transparent_pil_image(img, pad_px=pad_px)
        out = BytesIO()
        img.save(out, format="PNG")
        out.seek(0)
        return out
    except Exception:
        try:
            buf.seek(0)
        except Exception:
            pass
        return buf

def _fit_image_to_cell(
    img,
    row,
    col_zero_based,
    col_width_chars=42.0,
    row_height_points=50.0,
    pad_px=3,
    exact_fit=False,
):
    """Anchor an openpyxl image inside one worksheet cell.

    Parameters
    ----------
    exact_fit : bool
        False keeps the image's aspect ratio and centers it with padding.
        True stretches the image to exactly fill the target cell.
    """
    try:
        cell_w_px = max(24, int(float(col_width_chars) * 7 + 5))
        cell_h_px = max(20, int(float(row_height_points) * 96 / 72))

        if exact_fit:
            w_px = cell_w_px
            h_px = cell_h_px
            off_x = 0
            off_y = 0
        else:
            max_w = max(10, cell_w_px - 2 * pad_px)
            max_h = max(10, cell_h_px - 2 * pad_px)
            orig_w = max(1, float(getattr(img, "width", 1) or 1))
            orig_h = max(1, float(getattr(img, "height", 1) or 1))
            aspect = orig_w / orig_h
            w_px = min(max_w, int(max_h * aspect))
            h_px = int(w_px / aspect)
            if h_px > max_h:
                h_px = max_h
                w_px = int(h_px * aspect)
            w_px = max(8, int(w_px))
            h_px = max(8, int(h_px))
            off_x = pad_px + max(0, (max_w - w_px) // 2)
            off_y = pad_px + max(0, (max_h - h_px) // 2)

        img.width = w_px
        img.height = h_px
        anchor = OneCellAnchor()
        anchor.editAs = "oneCell"
        anchor._from = AnchorMarker(
            col=col_zero_based,
            row=row - 1,
            colOff=pixels_to_EMU(off_x),
            rowOff=pixels_to_EMU(off_y),
        )
        anchor.ext = XDRPositiveSize2D(cx=pixels_to_EMU(w_px), cy=pixels_to_EMU(h_px))
        img.anchor = anchor
    except Exception:
        # Safe fallback: still place the image in the requested cell.
        img.anchor = f"{get_column_letter(col_zero_based + 1)}{row}"
    return img

# ── write BBS sheet ─────────────────────────────────────────────────────────
def write_bbs(ws, bars, hdr, sketch_map=None, use_preview_sketches=False, fit_sketches_to_cell=False):
    ws.sheet_view.showGridLines = False

    for col, w in BBS_COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    for col in ("P", "R", "S"):
        ws.column_dimensions[col].hidden = True
    for ci in range(20, 47):
        ws.column_dimensions[get_column_letter(ci)].width = 13

    MED = Side(style="medium")
    MED_BORD = Border(left=MED, right=MED, top=MED, bottom=MED)
    BLANK_BORDER = Border()

    # ── header block rows 1-6 (matches the uploaded Armabeton-style target format) ──
    HDR_ROWS = [
        ("PROJECT NAME: ",   hdr.get("project_name", ""), "BBS REV.",          hdr.get("bbs_rev", "00")),
        ("BBS NAME : ",      hdr.get("bbs_name", ""),     "SHOP DRAWING REV.", hdr.get("sd_rev", "00")),
        ("BBS NO: ",         hdr.get("bbs_no", ""),       "OBJECT TYPE:",      hdr.get("object_type", "")),
        ("SHOP DRAWING NO.", hdr.get("sd_no", ""),        "PREPARED BY:",      hdr.get("prepared_by", "")),
        ("DETAILING OFFICE", hdr.get("office", ""),       "CHECKED BY:",       hdr.get("checked_by", "")),
        ("", "", "", ""),
    ]

    for ri, (lbl1, val1, lbl2, val2) in enumerate(HDR_ROWS, 1):
        ws.row_dimensions[ri].height = 15 if ri < 6 else 15.75
        ws.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=11)  # B:K
        ws.merge_cells(start_row=ri, start_column=14, end_row=ri, end_column=15) # N:O
        vals = {1: lbl1 or None, 2: val1 or None, 14: lbl2 or None, 17: val2 or None}
        for ci in range(1, 18):
            c = ws.cell(row=ri, column=ci, value=vals.get(ci))
            c.font = Font(**FONT_C11)
            c.alignment = LEFT if ci in (1, 2, 14) else CTR
            c.border = ALL_T
            if ci == 17:
                c.number_format = "@"
        for ci in range(20, 47):
            c = ws.cell(row=ri, column=ci)
            c.fill = FILL_DARK
            c.border = BLANK_BORDER

    # Divider row
    ws.row_dimensions[7].height = 13.5
    ws.merge_cells("A7:Q7")
    for ci in range(1, 18):
        c = ws.cell(row=7, column=ci)
        c.alignment = CTR
        c.border = Border(top=MED, bottom=MED)
    for ci in range(20, 47):
        ws.cell(row=7, column=ci).fill = FILL_DARK

    # ── row 8: column headers ──
    ws.row_dimensions[8].height = 86.25
    for ci, label in enumerate(BBS_COL_HDRS, 1):
        cell = ws.cell(row=8, column=ci, value=label)
        cell.font = Font(**FONT_A9B)
        cell.fill = FILL_GREY
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = MED_BORD
    ws.cell(row=8, column=19).border = BLANK_BORDER
    for ci in range(20, 47):
        c = ws.cell(row=8, column=ci)
        c.fill = FILL_DARK
        c.border = BLANK_BORDER

    try:
        from PIL import Image as _PILImage; has_pil = True
    except ImportError:
        has_pil = False

    # ── data rows ──
    ROW_H = 50.0
    for row_off, row_data in bars.iterrows():
        xrow = 9 + row_off
        ws.row_dimensions[xrow].height = ROW_H
        dims = row_data["dims"] if isinstance(row_data["dims"], dict) else {}

        data = [
            row_data["idx"], row_data["bar_type"], row_data["nb_elem"], row_data["nb_steel"],
            row_data["nb_total"], row_data["mandrin"], row_data.get("shape_code"),
            dims.get("A"), dims.get("B"), dims.get("C"), dims.get("D"), dims.get("E"), dims.get("F"), dims.get("G"),
            row_data["total_len"], None, None, None, None,
        ]
        for ci, val in enumerate(data, 1):
            try:
                val2 = float(val) if val not in (None, "", "nan") else None
                if val2 is not None and val2 == int(val2):
                    val2 = int(val2)
                val = val2
            except Exception:
                pass
            c = ws.cell(row=xrow, column=ci, value=val)
            c.font = Font(**FONT_A10R) if ci in (1, 2) else Font(**FONT_A10)
            c.alignment = CTR
            c.border = ALL_T if ci <= 18 else BLANK_BORDER
        for ci in range(20, 47):
            c = ws.cell(row=xrow, column=ci)
            c.fill = FILL_DARK
            c.border = BLANK_BORDER

        if has_pil and (use_preview_sketches or not sketch_map):
            # Generate sketches when requested, or when no source sketches are available. Sketch column is Q.
            buf = draw_sketch(dims, total_len=row_data.get("total_len"))
            if buf:
                # Preview-toggle sketches use a generated transparent canvas.
                # Trim that canvas before anchoring. The checkbox/CLI option below
                # chooses between centered aspect-ratio fit and exact full-cell fit.
                buf = _trim_preview_sketch_buffer(buf, pad_px=10)
                img = XLImage(buf)
                img = _fit_image_to_cell(
                    img, xrow, 16,
                    col_width_chars=BBS_COL_WIDTHS.get("Q", 42.0),
                    row_height_points=ROW_H,
                    pad_px=3,
                    exact_fit=fit_sketches_to_cell,
                )
                ws.add_image(img)

    ws.freeze_panes = "A9"


# ── write TONNAGE ────────────────────────────────────────────────────────────
def write_tonnage(ws, bars, unit_weights=None):
    """Write the Armabeton-style TONNAGE sheet.

    The layout/format matches the uploaded target sheet, while derived columns
    are kept as formulas. Formula cached values are inserted after save by
    cache_tonnage_formula_values() so viewers can show values immediately.
    """
    unit_weights = unit_weights or {}

    # The uploaded target TONNAGE sheet has no tab colour. Keep gridlines visible
    # and use exact #D9D9D9 background on the title/header areas.
    ws.sheet_properties.tabColor = None
    ws.sheet_view.showGridLines = True

    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.column_dimensions["H"].width = 12

    FILL_GREY = PatternFill("solid", fgColor="FFD9D9D9")
    FONT_TITLE = Font(name="Calibri", size=20, bold=True, color="FF0070C0")
    FONT_ELEM  = Font(name="Calibri", size=14)
    FONT_STEEL = Font(name="Calibri", size=12, bold=True, italic=True, color="FF339966")
    FONT_HDR_B = Font(name="Calibri", size=11, bold=True, color="FF000000")
    FONT_HDR_N = Font(name="Calibri", size=11, color="FF000000")
    FONT_NORM  = Font(name="Arial",   size=10)
    FONT_TOTAL_LBL = Font(name="Arial", size=10, bold=True, italic=True, color="FFFF0000")
    FONT_TOTAL = Font(name="Arial", size=10, bold=True, color="FFFF0000")

    ALIGN_CTR  = Alignment(horizontal="center", vertical="center")
    ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
    THIN = Side(style="thin")
    MED  = Side(style="medium")
    BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BORD_MED = Border(left=MED, right=MED, top=MED, bottom=MED)

    def put(row, col, value=None, font=FONT_NORM, fill=None, align=ALIGN_CTR, border=BORD, fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font
        c.alignment = align
        c.border = border
        if fill:
            c.fill = fill
        if fmt:
            c.number_format = fmt
        return c

    # Row geometry from the uploaded target workbook.
    ws.row_dimensions[3].height = 13.5
    ws.row_dimensions[4].height = 27
    ws.row_dimensions[5].height = 19.5
    ws.row_dimensions[7].height = 15.75
    for r in (9, 10, 11):
        ws.row_dimensions[r].height = 15

    # Title block.
    ws.merge_cells("B4:I4")
    put(4, 2, "WEIGHT OF STEEL", FONT_TITLE, FILL_GREY, ALIGN_CTR, BORD_MED)
    for c in range(3, 10):
        ws.cell(row=4, column=c).fill = FILL_GREY
        ws.cell(row=4, column=c).border = BORD_MED

    ws.merge_cells("B5:I5")
    put(5, 2, "Number of elements: 1", FONT_ELEM, None, ALIGN_CTR, BORD_MED)
    for c in range(3, 10):
        ws.cell(row=5, column=c).border = BORD_MED

    # Empty framed area between title and table.  The steel grade line is kept as-is.
    for r in (6, 7, 8):
        for c in range(2, 10):
            ws.cell(row=r, column=c).border = Border(
                left=MED if c == 2 else None,
                right=MED if c == 9 else None,
                bottom=MED if r == 8 else None,
            )
    put(7, 2, "Steels HA -Fe E 420:", FONT_STEEL, None, ALIGN_LEFT, Border(left=MED))

    # Three-row table header.
    headers = {
        9:  [None, None, "Unshaped", "Faceted", "Total ", "Unit", "Unformed", "Shaped", "Total "],
        10: [None, "Diameter", "Length", "Length", "Length", "Weight", "Weight", "Weight", "Weight"],
        11: [None, "(mm)", "(m)", "(m)", "(m)", "(kg/m)", "(kg)", "(kg)", "(kg)"],
    }
    for row, vals in headers.items():
        for ci in range(2, 10):
            val = vals[ci-1]
            font = FONT_HDR_B if row in (9, 10) else FONT_HDR_N
            border = Border(
                left=MED if ci == 2 else THIN,
                right=MED if ci == 9 else THIN,
                top=MED if row == 9 else THIN,
                bottom=MED if row == 11 else THIN,
            )
            put(row, ci, val, font, FILL_GREY, ALIGN_CTR, border)

    # Build data by diameter — split unshaped vs faceted based on dims count.
    by_dia_unshaped = {}
    by_dia_faceted  = {}
    for _, r in bars.iterrows():
        d = int(r["diameter"]) if r["diameter"] else 0
        tl = float(r["total_len"]) / 1000.0 * int(r["nb_total"])
        dims = r["dims"] if isinstance(r["dims"], dict) else {}
        if len(dims) <= 1:
            by_dia_unshaped[d] = by_dia_unshaped.get(d, 0.0) + tl
        else:
            by_dia_faceted[d] = by_dia_faceted.get(d, 0.0) + tl

    source_dias = [d for d in sorted(unit_weights) if d in {6,8,10,12,16,20,25,28,32,40}]
    bar_dias = sorted(set(by_dia_unshaped) | set(by_dia_faceted))
    diameters = [d for d in source_dias if d in set(bar_dias) or d in {10,12,16,20,25,28,32}]
    for d in bar_dias:
        if d and d not in diameters:
            diameters.append(d)
    if not diameters:
        diameters = [10, 12, 16, 20, 25, 28, 32]

    DATA_START = 12
    data_rows = []
    for d_off, d in enumerate(diameters):
        row = DATA_START + d_off
        u_m = round(by_dia_unshaped.get(d, 0.0), 3)
        f_m = round(by_dia_faceted.get(d,  0.0), 3)
        total_m = round(u_m + f_m, 3)
        wperm = unit_weights.get(d)
        if wperm is None:
            # Density-based fallback, only used when the source sheet does not provide the diameter.
            wperm = round(float(d) * float(d) * 0.006165, 3)
        wperm = float(wperm)
        data_rows.append(row)

        row_values = {
            2: d,
            3: u_m if total_m else 0,
            4: f_m if total_m else 0,
            5: f"=C{row}+D{row}",
            6: wperm,
            7: f"=C{row}*F{row}",
            8: f"=D{row}*F{row}",
            9: f"=G{row}+H{row}",
        }
        for ci in range(2, 10):
            border = Border(
                left=MED if ci == 2 else THIN,
                right=MED if ci == 9 else THIN,
                top=THIN,
                bottom=MED if row == DATA_START + len(diameters) - 1 else THIN,
            )
            # Match the target sheet: detail rows use General, totals use integer display.
            put(row, ci, row_values.get(ci), FONT_NORM, None, ALIGN_CTR, border, "General")

    # Totals row with formulas and separated blocks like the target.
    tot_row = DATA_START + len(diameters) + 2
    put(tot_row, 2, "Totals:", FONT_TOTAL_LBL, None, ALIGN_LEFT, Border())
    for ci in range(3, 10):
        if ci == 6:
            ws.cell(row=tot_row, column=ci).value = None
            ws.cell(row=tot_row, column=ci).font = FONT_TOTAL
            ws.cell(row=tot_row, column=ci).number_format = "General"
            continue
        col_l = get_column_letter(ci)
        put(tot_row, ci, f"=SUM({col_l}{data_rows[0]}:{col_l}{data_rows[-1]})",
            FONT_TOTAL, None, ALIGN_CTR, BORD, "0")


# ── extract sketches from source xls(x) ─────────────────────────────────────
def extract_source_sketches(src_path):
    """Extract WMF sketch bytes and source display aspect ratios.

    Returns {src_row: (wmf_bytes, wmf_name, aspect_ratio, orig_cx, orig_cy)}.
    The aspect ratio and original source drawing extents are taken from the
    source drawing xfrm extent when available. The injected sketch is kept at
    its original displayed size when it fits, then centered in the destination
    cell.
    """
    import zipfile
    from xml.etree import ElementTree as ET

    result = {}
    try:
        with zipfile.ZipFile(src_path) as z:
            all_files = z.namelist()
            if not any(f.endswith('.wmf') for f in all_files):
                return result

            # Find Nomenclature sheet index.
            wb_root = ET.fromstring(z.read('xl/workbook.xml').decode())
            ns_wb = {'wb': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            nom_sheet_id = 3  # fallback
            for i, s in enumerate(wb_root.findall('.//wb:sheet', ns_wb), 1):
                if 'nomenclature' in s.attrib.get('name', '').lower():
                    nom_sheet_id = i
                    break

            # Find drawing relationship for that sheet.
            rel_path = f'xl/worksheets/_rels/sheet{nom_sheet_id}.xml.rels'
            if rel_path not in all_files:
                return result
            drawing_target = None
            for rel in ET.fromstring(z.read(rel_path).decode()):
                if 'drawing' in rel.attrib.get('Type', ''):
                    drawing_target = rel.attrib['Target'].replace('../', 'xl/')
                    break
            if not drawing_target or drawing_target not in all_files:
                return result

            # Build rId -> WMF name.
            draw_rel_path = drawing_target.replace('drawings/', 'drawings/_rels/') + '.rels'
            if draw_rel_path not in all_files:
                return result
            rid_to_wmf = {}
            for rel in ET.fromstring(z.read(draw_rel_path).decode()):
                rid_to_wmf[rel.attrib['Id']] = rel.attrib['Target'].replace('../media/', '')

            ns = {
                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                'a':   'http://schemas.openxmlformats.org/drawingml/2006/main',
                'r':   'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
            }
            root = ET.fromstring(z.read(drawing_target).decode())

            for anchor in list(root.findall('.//xdr:twoCellAnchor', ns)) + list(root.findall('.//xdr:oneCellAnchor', ns)):
                from_el = anchor.find('xdr:from', ns)
                if from_el is None:
                    continue
                row_el = from_el.find('xdr:row', ns)
                if row_el is None:
                    continue
                row = int(row_el.text)

                blip = anchor.find('.//a:blip', ns)
                if blip is None:
                    continue
                rid = blip.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                wmf_name = rid_to_wmf.get(rid)
                if not wmf_name:
                    continue
                wmf_path = f'xl/media/{wmf_name}'
                if wmf_path not in all_files:
                    continue

                aspect = None
                orig_cx = orig_cy = None
                ext_el = anchor.find('.//a:xfrm/a:ext', ns)
                if ext_el is None:
                    ext_el = anchor.find('xdr:ext', ns)
                if ext_el is not None:
                    try:
                        cx = int(float(ext_el.attrib.get('cx', 0)))
                        cy = int(float(ext_el.attrib.get('cy', 0)))
                        if cx > 0 and cy > 0:
                            orig_cx, orig_cy = cx, cy
                            aspect = cx / cy
                    except Exception:
                        aspect = None
                        orig_cx = orig_cy = None
                if not aspect or aspect <= 0:
                    aspect = 600.0 / 220.0

                result[row] = (z.read(wmf_path), wmf_name, aspect, orig_cx, orig_cy)
    except Exception as e:
        print(f"  [sketch extract warning: {e}]")
    return result


def inject_wmf_sketches(dst_path, bars, sketch_map, col_idx=16, fit_sketches_to_cell=False):
    """Post-process saved xlsx to inject WMF sketches into the SKETCH column.

    By default, source WMF images keep their aspect ratio and are centered in the
    Q-column cell.  If fit_sketches_to_cell=True, the WMF extents are set to the
    full cell dimensions so the sketch exactly fills the cell.
    """
    import zipfile, shutil, os

    if not sketch_map:
        return

    tmp_path = dst_path + '.tmp'
    shutil.copy2(dst_path, tmp_path)

    # Build bar -> (wmf bytes, name, aspect, source extents) using src_row.
    bar_sketches = []  # list of (xrow_0based, wmf_bytes, wmf_name, aspect, orig_cx, orig_cy)
    xrow = 8  # BBS data starts at row 9 (0-based = 8), each bar = 1 row
    bar_iter = bars.to_dict('records') if hasattr(bars, 'to_dict') else bars
    for bar in bar_iter:
        src_row = bar.get('src_row', -1)
        if src_row in sketch_map:
            entry = sketch_map[src_row]
            wmf_bytes, wmf_name = entry[0], entry[1]
            aspect = entry[2] if len(entry) > 2 and entry[2] else 600.0 / 220.0
            orig_cx = entry[3] if len(entry) > 3 else None
            orig_cy = entry[4] if len(entry) > 4 else None
            try:
                aspect = float(aspect)
                if aspect <= 0:
                    aspect = 600.0 / 220.0
            except Exception:
                aspect = 600.0 / 220.0
            bar_sketches.append((xrow, wmf_bytes, wmf_name, aspect, orig_cx, orig_cy))
        xrow += 1

    if not bar_sketches:
        os.remove(tmp_path)
        return

    with zipfile.ZipFile(tmp_path, 'r') as zin:
        file_data = {n: zin.read(n) for n in zin.namelist()}

    wmf_store = {}
    row_to_safe = {}
    row_to_aspect = {}
    for xrow, wmf_bytes, wmf_name, aspect, orig_cx, orig_cy in bar_sketches:
        safe = f'sketch_{xrow}.wmf'
        wmf_store[safe] = wmf_bytes
        row_to_safe[xrow] = safe
        row_to_aspect[xrow] = aspect

    for safe, wmf_bytes in wmf_store.items():
        file_data[f'xl/media/{safe}'] = wmf_bytes

    anchors_xml = []
    rels_entries = []

    # Cell dimensions in EMU.  Keep in sync with BBS_COL_WIDTHS["Q"] and ROW_H.
    COL_WIDTH_CHARS = 42.0
    ROW_HEIGHT_PT = 50.0
    CELL_W = int(COL_WIDTH_CHARS * 7 * 9144)
    CELL_H = int(ROW_HEIGHT_PT * 12700)
    PAD = int(0.06 * min(CELL_W, CELL_H))
    MAX_W = max(1, CELL_W - PAD * 2)
    MAX_H = max(1, CELL_H - PAD * 2)

    for i, (xrow, wmf_bytes, wmf_name, aspect, orig_cx, orig_cy) in enumerate(bar_sketches):
        rid = f'rId{i+1}'
        safe = row_to_safe[xrow]
        rels_entries.append(
            f'<Relationship Id="{rid}" '
            f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
            f'Target="../media/{safe}"/>'
        )

        if fit_sketches_to_cell:
            # Exact-fit mode deliberately uses the full cell extents. This can
            # change the sketch aspect ratio, but it makes the sketch occupy the
            # whole SKETCH cell as requested.
            img_w = CELL_W
            img_h = CELL_H
            off_x = 0
            off_y = 0
        else:
            # Centered mode keeps the source display aspect ratio. Keep original
            # displayed size when it fits, scale down only if needed, and center.
            try:
                img_w = int(orig_cx or 0)
                img_h = int(orig_cy or 0)
            except Exception:
                img_w = img_h = 0
            if img_w <= 0 or img_h <= 0:
                img_h = min(MAX_H, 460375)
                img_w = int(round(img_h * aspect))
            scale = min(1.0, MAX_W / img_w, MAX_H / img_h)
            img_w = max(1, int(round(img_w * scale)))
            img_h = max(1, int(round(img_h * scale)))
            off_x = PAD + max(0, (MAX_W - img_w) // 2)
            off_y = PAD + max(0, (MAX_H - img_h) // 2)

        anchors_xml.append(
            f'<xdr:oneCellAnchor editAs="oneCell">'
            f'<xdr:from><xdr:col>{col_idx}</xdr:col><xdr:colOff>{off_x}</xdr:colOff>'
            f'<xdr:row>{xrow}</xdr:row><xdr:rowOff>{off_y}</xdr:rowOff></xdr:from>'
            f'<xdr:ext cx="{img_w}" cy="{img_h}"/>'
            f'<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="{i+1}" name="Sketch{i+1}"/>'
            + (f'<xdr:cNvPicPr/>' if fit_sketches_to_cell else f'<xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr>')
            + f'</xdr:nvPicPr>'
            f'<xdr:blipFill><a:blip r:embed="{rid}"/><a:srcRect/><a:stretch><a:fillRect/></a:stretch></xdr:blipFill>'
            f'<xdr:spPr><a:xfrm><a:off x="0" y="0"/>'
            f'<a:ext cx="{img_w}" cy="{img_h}"/></a:xfrm>'
            f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f'<a:ln w="0"><a:noFill/></a:ln></xdr:spPr>'
            f'</xdr:pic><xdr:clientData/></xdr:oneCellAnchor>'
        )

    ns_xdr = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
    ns_a   = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    ns_r   = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

    drawing_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<xdr:wsDr xmlns:xdr="{ns_xdr}" xmlns:a="{ns_a}" xmlns:r="{ns_r}">' +
        ''.join(anchors_xml) + '</xdr:wsDr>'
    )

    drawing_path = 'xl/drawings/drawing_bbs.xml'
    drawing_rels_path = 'xl/drawings/_rels/drawing_bbs.xml.rels'
    drawing_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">' +
        ''.join(rels_entries) + '</Relationships>'
    )
    file_data[drawing_path] = drawing_xml.encode()
    file_data[drawing_rels_path] = drawing_rels_xml.encode()

    sheet1_path = 'xl/worksheets/sheet1.xml'
    if sheet1_path in file_data:
        s1 = file_data[sheet1_path].decode()
        if '<drawing' not in s1:
            s1 = s1.replace(
                '</worksheet>',
                '<drawing xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId_draw"/></worksheet>'
            )
            file_data[sheet1_path] = s1.encode()

    sheet1_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId_draw" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
        'Target="../drawings/drawing_bbs.xml"/>'
        '</Relationships>'
    )
    file_data['xl/worksheets/_rels/sheet1.xml.rels'] = sheet1_rels.encode()

    ct_path = '[Content_Types].xml'
    if ct_path in file_data:
        ct = file_data[ct_path].decode()
        if 'Extension="wmf"' not in ct:
            ct = ct.replace('</Types>', '<Default Extension="wmf" ContentType="image/x-wmf"/></Types>')
        if 'drawing_bbs' not in ct:
            ct = ct.replace(
                '</Types>',
                '<Override PartName="/xl/drawings/drawing_bbs.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/></Types>'
            )
        file_data[ct_path] = ct.encode()

    with zipfile.ZipFile(dst_path, 'w', zipfile.ZIP_STORED) as zout:
        for name, data in file_data.items():
            zout.writestr(name, data)

    os.remove(tmp_path)
    print(f"  [injected {len(bar_sketches)} WMF sketches; exact_fit={fit_sketches_to_cell}]")


# ── formula cache injection ──────────────────────────────────────────────────
def cache_tonnage_formula_values(dst_path):
    """Insert cached numeric values for TONNAGE formulas while keeping formulas.

    openpyxl writes formulas but not their cached results.  This post-process makes
    lightweight previews/data-only readers show TONNAGE values immediately without
    removing formulas for Excel.
    """
    import os, re, zipfile, shutil
    from xml.etree import ElementTree as ET

    if not os.path.isfile(dst_path):
        return

    NS_MAIN = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    NS_REL  = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    NS_PKG  = 'http://schemas.openxmlformats.org/package/2006/relationships'
    ET.register_namespace('', NS_MAIN)
    ET.register_namespace('r', NS_REL)

    def col_to_num(col):
        n = 0
        for ch in col:
            n = n * 26 + ord(ch.upper()) - 64
        return n

    def a1_to_rc(addr):
        m = re.match(r'^([A-Z]+)(\d+)$', addr.upper())
        if not m:
            return None
        return int(m.group(2)), col_to_num(m.group(1))

    def num_to_col(n):
        s = ''
        while n:
            n, rem = divmod(n - 1, 26)
            s = chr(65 + rem) + s
        return s

    def fmt_num(v):
        try:
            v = float(v)
        except Exception:
            return None
        if abs(v - round(v)) < 1e-10:
            return str(int(round(v)))
        return f'{v:.12f}'.rstrip('0').rstrip('.')

    def getv(addr, values):
        return values.get(addr.upper(), 0.0)

    def eval_formula(formula, values):
        f = (formula or '').strip()
        if f.startswith('='):
            f = f[1:]
        f = f.replace('$', '')
        m = re.fullmatch(r'([A-Z]+\d+)\+([A-Z]+\d+)', f, re.I)
        if m:
            return getv(m.group(1), values) + getv(m.group(2), values)
        m = re.fullmatch(r'([A-Z]+\d+)\*([A-Z]+\d+)', f, re.I)
        if m:
            return getv(m.group(1), values) * getv(m.group(2), values)
        m = re.fullmatch(r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', f, re.I)
        if m:
            c1, r1, c2, r2 = m.group(1).upper(), int(m.group(2)), m.group(3).upper(), int(m.group(4))
            n1, n2 = col_to_num(c1), col_to_num(c2)
            total = 0.0
            for rr in range(min(r1, r2), max(r1, r2) + 1):
                for cc in range(min(n1, n2), max(n1, n2) + 1):
                    total += getv(f'{num_to_col(cc)}{rr}', values)
            return total
        return None

    tmp_path = dst_path + '.formula_cache_tmp'
    shutil.copy2(dst_path, tmp_path)
    try:
        with zipfile.ZipFile(tmp_path, 'r') as zin:
            file_data = {name: zin.read(name) for name in zin.namelist()}

        wb_xml = ET.fromstring(file_data['xl/workbook.xml'])
        wb_rels = ET.fromstring(file_data['xl/_rels/workbook.xml.rels'])
        rid = None
        for sheet in wb_xml.findall(f'.//{{{NS_MAIN}}}sheet'):
            if sheet.attrib.get('name') == 'TONNAGE':
                rid = sheet.attrib.get(f'{{{NS_REL}}}id')
                break
        if not rid:
            return
        target = None
        for rel in wb_rels.findall(f'{{{NS_PKG}}}Relationship'):
            if rel.attrib.get('Id') == rid:
                target = rel.attrib.get('Target')
                break
        if not target:
            return
        if target.startswith('/'):
            sheet_path = target.lstrip('/')
        elif target.startswith('xl/'):
            sheet_path = target
        else:
            sheet_path = 'xl/' + target
        if sheet_path not in file_data:
            return

        root = ET.fromstring(file_data[sheet_path])
        values = {}
        formulas = {}
        for c in root.findall(f'.//{{{NS_MAIN}}}c'):
            addr = c.attrib.get('r')
            if not addr:
                continue
            f_el = c.find(f'{{{NS_MAIN}}}f')
            v_el = c.find(f'{{{NS_MAIN}}}v')
            if v_el is not None and v_el.text not in (None, ''):
                try:
                    values[addr.upper()] = float(v_el.text)
                except Exception:
                    pass
            if f_el is not None and f_el.text:
                formulas[addr.upper()] = f_el.text

        # Evaluate simple arithmetic and SUM formulas iteratively, because some formulas
        # depend on other formula cells.
        for _ in range(8):
            changed = False
            for addr, formula in formulas.items():
                val = eval_formula(formula, values)
                if val is None:
                    continue
                old = values.get(addr)
                values[addr] = val
                if old is None or abs(old - val) > 1e-10:
                    changed = True
            if not changed:
                break

        for c in root.findall(f'.//{{{NS_MAIN}}}c'):
            addr = c.attrib.get('r', '').upper()
            if addr not in formulas or addr not in values:
                continue
            v_el = c.find(f'{{{NS_MAIN}}}v')
            if v_el is None:
                v_el = ET.SubElement(c, f'{{{NS_MAIN}}}v')
            v_el.text = fmt_num(values[addr])

        file_data[sheet_path] = ET.tostring(root, encoding='utf-8', xml_declaration=True)
        with zipfile.ZipFile(dst_path, 'w', zipfile.ZIP_STORED) as zout:
            for name, data in file_data.items():
                zout.writestr(name, data)
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass


# ── convert ──────────────────────────────────────────────────────────────────
def convert(src, dst, hdr, progress_cb=None, use_preview_sketches_for_excel=False, fit_sketches_to_cell=False):
    def pct(p, m=""):
        if progress_cb: progress_cb(p, m)
    pct(5,  "Reading source…")
    bars = parse_nomenclature(src)
    pct(18, f"Parsed {len(bars)} bars")
    pct(22, "Reading exact unit weights…")
    unit_weights = parse_unit_weights(src)
    pct(25, "Reading couplers…")
    coupler_data = parse_couplers(src)
    pct(28, "Extracting sketches…")
    sketch_map = extract_source_sketches(src)
    if use_preview_sketches_for_excel:
        mode = "exact full-cell" if fit_sketches_to_cell else "centered aspect-ratio"
        pct(43, f"Using generated preview sketches for Excel output ({mode})…")
    elif sketch_map:
        pct(43, f"Found {len(sketch_map)} source sketches…")
    else:
        pct(43, "No source sketches found — using generated sketches…")

    wb = Workbook(); ws_bbs = wb.active; ws_bbs.title = "BAR BENDING SCHEDULE"
    ws_bbs.sheet_properties.tabColor = "92D050"
    pct(50, "Writing BBS sheet…")
    write_bbs(ws_bbs, bars, hdr,
              sketch_map=None if use_preview_sketches_for_excel else sketch_map,
              use_preview_sketches=use_preview_sketches_for_excel,
              fit_sketches_to_cell=fit_sketches_to_cell)
    pct(65, "Writing Tonnage sheet…")
    ws_ton = wb.create_sheet("TONNAGE"); write_tonnage(ws_ton, bars, unit_weights=unit_weights)
    if coupler_data:
        pct(70, "Writing couplers block…")
        write_couplers_on_tonnage(ws_ton, coupler_data)

    # No support sheets are copied or generated. Couplers are added on TONNAGE only when present.

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    pct(85, "Saving…"); wb.save(dst)
    if sketch_map and not use_preview_sketches_for_excel:
        pct(90, "Injecting source sketches…")
        inject_wmf_sketches(dst, bars, sketch_map, fit_sketches_to_cell=fit_sketches_to_cell)
    pct(96, "Caching tonnage formula values…")
    cache_tonnage_formula_values(dst)
    pct(100, "Done!")


# ── batch convert ────────────────────────────────────────────────────────────
def batch_convert(jobs, hdr_defaults=None, progress_cb=None, log_cb=None, use_preview_sketches_for_excel=False, fit_sketches_to_cell=False):
    """
    Convert multiple files.

    Parameters
    ----------
    jobs : list of dict
        Each dict must have:
          'src'  – source .xls/.xlsx path
          'dst'  – output .xlsx path  (auto-derived if absent/empty)
        Optionally:
          'hdr'  – per-file header overrides (merged with hdr_defaults)
    hdr_defaults : dict, optional
        Header values applied to every job unless overridden per-job.
    progress_cb : callable(job_index, n_jobs, file_pct, message), optional
        Called frequently with overall + per-file progress info.
    log_cb : callable(message), optional
        Called with one-line status strings (success / skip / error per file).
    use_preview_sketches_for_excel : bool, optional
        If True, generated preview sketches are written into Excel instead of source WMF sketches.
    fit_sketches_to_cell : bool, optional
        If True, sketches are stretched to exactly fill the SKETCH cell.

    Returns
    -------
    results : list of dict  {'src', 'dst', 'ok': bool, 'error': str|None}
    """
    hdr_defaults = hdr_defaults or {}
    results = []
    n = len(jobs)

    for i, job in enumerate(jobs):
        src = job.get("src", "").strip()
        dst = job.get("dst", "").strip()
        if not dst:
            base = os.path.splitext(src)[0]
            dst = base + "_BBS.xlsx"
        hdr = {**hdr_defaults, **job.get("hdr", {})}
        job_use_preview = job.get("use_preview_sketches_for_excel", use_preview_sketches_for_excel)
        job_fit_to_cell = job.get("fit_sketches_to_cell", fit_sketches_to_cell)

        if not os.path.isfile(src):
            msg = f"[{i+1}/{n}] SKIP (not found): {src}"
            if log_cb: log_cb(msg)
            results.append({"src": src, "dst": dst, "ok": False, "error": "File not found"})
            continue

        if log_cb: log_cb(f"[{i+1}/{n}] Converting: {os.path.basename(src)} …")

        def _pct(p, m, _i=i):
            if progress_cb:
                progress_cb(_i, n, p, m)

        try:
            convert(src, dst, hdr, progress_cb=_pct,
                    use_preview_sketches_for_excel=job_use_preview,
                    fit_sketches_to_cell=job_fit_to_cell)
            msg = f"[{i+1}/{n}] ✔  {os.path.basename(src)}  →  {os.path.basename(dst)}"
            if log_cb: log_cb(msg)
            results.append({"src": src, "dst": dst, "ok": True, "error": None})
        except Exception as e:
            msg = f"[{i+1}/{n}] ✘  {os.path.basename(src)}: {e}"
            if log_cb: log_cb(msg)
            results.append({"src": src, "dst": dst, "ok": False, "error": str(e)})

    if log_cb:
        n_ok  = sum(r["ok"] for r in results)
        n_err = n - n_ok
        log_cb(f"\nBatch complete — {n_ok} succeeded, {n_err} failed.")
    return results


# ── GUI ──────────────────────────────────────────────────────────────────────
def launch_gui():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    import json, math, ctypes

    def get_config_path():
        """Return a persistent config path that also works after PyInstaller EXE build.

        In a bundled EXE, __file__ can point to a temporary extraction folder, so
        saving beside __file__ may disappear after closing the app. Store user
        settings in AppData instead, and fall back to the script folder only if
        AppData is not available.
        """
        app_name = "BBS Transformer"
        file_name = "bbs_transformer_config.json"

        if sys.platform == "win32":
            base = (os.environ.get("APPDATA") or
                    os.environ.get("LOCALAPPDATA") or
                    os.path.expanduser("~"))
            config_dir = os.path.join(base, app_name)
        else:
            config_dir = os.path.join(os.path.expanduser("~"), ".config", "bbs_transformer")

        try:
            os.makedirs(config_dir, exist_ok=True)
            return os.path.join(config_dir, file_name)
        except Exception:
            return os.path.join(os.path.dirname(os.path.abspath(sys.argv[0])), file_name)

    CONFIG_PATH = get_config_path()

    def migrate_old_config_if_needed():
        """Move old beside-script/temporary config into the persistent path once."""
        if os.path.exists(CONFIG_PATH):
            return
        old_locations = []
        try:
            old_locations.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), "bbs_transformer_config.json"))
        except Exception:
            pass
        try:
            old_locations.append(os.path.join(os.path.dirname(os.path.abspath(sys.argv[0])), "bbs_transformer_config.json"))
        except Exception:
            pass

        for old_path in old_locations:
            try:
                if old_path and old_path != CONFIG_PATH and os.path.exists(old_path):
                    with open(old_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    os.makedirs(os.path.dirname(CONFIG_PATH), exist_ok=True)
                    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                        json.dump(data, f, indent=2)
                    return
            except Exception:
                pass

    migrate_old_config_if_needed()

    def load_config():
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}

    def save_config(data):
        try:
            os.makedirs(os.path.dirname(CONFIG_PATH), exist_ok=True)
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
        except Exception as e:
            try:
                messagebox.showwarning("Settings not saved", f"Could not save settings:\n{e}\n\nPath:\n{CONFIG_PATH}")
            except Exception:
                pass

    # ── Windows 11 Mica glass + rounded corners ────────────────────────────
    def set_win11_glass(win):
        """Best-effort Win11 Mica/Acrylic backdrop + rounded corners. No-op elsewhere."""
        try:
            if sys.platform != "win32": return
            win.update_idletasks()
            hwnd = ctypes.windll.user32.GetParent(win.winfo_id()) or win.winfo_id()
            dwm = ctypes.windll.dwmapi

            # DWMWA_WINDOW_CORNER_PREFERENCE = 33, DWMWCP_ROUND = 2
            dwm.DwmSetWindowAttribute(hwnd, 33, ctypes.byref(ctypes.c_int(2)), 4)

            # DWMWA_SYSTEMBACKDROP_TYPE = 38. 3 = Acrylic, 2 = Mica.
            # Acrylic gives the closest "glass" feel; if unavailable Windows
            # simply ignores it, and the Tk light-glass palette remains.
            backdrop = ctypes.c_int(3)
            if dwm.DwmSetWindowAttribute(hwnd, 38, ctypes.byref(backdrop), 4) != 0:
                backdrop = ctypes.c_int(2)
                dwm.DwmSetWindowAttribute(hwnd, 38, ctypes.byref(backdrop), 4)

            # DWMWA_USE_IMMERSIVE_DARK_MODE = 20; match the dark-blue UI.
            dwm.DwmSetWindowAttribute(hwnd, 20, ctypes.byref(ctypes.c_int(1)), 4)
        except Exception:
            pass

    # ── RoundedCard / RoundedButton ──────────────────────────────────────────
    class RoundedCard(tk.Frame):
        """Clean Win11-style rounded card.

        The outer frame uses the parent's background.  A rounded shape is drawn
        behind an inset inner frame.  Children go into .inner, which prevents
        the jagged/saw-tooth edges that happen when labels are placed directly
        on a transparent-looking rounded frame.
        """
        def __init__(self, parent, fill, outline, radius=18, pad=8, **kw):
            kw.pop("bg", None)
            try:
                parent_bg = parent.cget("bg")
            except Exception:
                parent_bg = C.get("bg", fill) if "C" in globals() else fill
            super().__init__(parent, bg=parent_bg, **kw)
            self._fill = fill
            self._outline = outline
            self._radius = radius
            self._pad = pad
            self._parent_bg = parent_bg
            # Keep the rectangular inner Frame slightly away from the edge so
            # the canvas-painted rounded corner remains visible instead of
            # looking like a square tab.
            self._inner_inset = max(6, min(10, radius // 3))

            self._cv = tk.Canvas(self, bg=parent_bg, highlightthickness=0, bd=0)
            self._cv.place(x=0, y=0, relwidth=1, relheight=1)

            self.inner = tk.Frame(self, bg=fill, padx=pad, pady=pad)
            self.inner.pack(fill="both", expand=True,
                            padx=self._inner_inset, pady=self._inner_inset)
            self.bind("<Configure>", self._redraw)

        def _rrect(self, x1, y1, x2, y2, r, **kw):
            r = max(4, min(r, int((x2-x1)/2), int((y2-y1)/2)))
            pts = [x1+r,y1, x2-r,y1, x2,y1, x2,y1+r,
                   x2,y2-r, x2,y2, x2-r,y2, x1+r,y2,
                   x1,y2, x1,y2-r, x1,y1+r, x1,y1]
            self._cv.create_polygon(pts, smooth=True, splinesteps=24, tags="bg", **kw)

        def _redraw(self, e=None):
            w = max(4, self.winfo_width() if e is None else e.width)
            h = max(4, self.winfo_height() if e is None else e.height)
            self._cv.config(width=w, height=h, bg=self._parent_bg)
            self._cv.delete("bg")
            self._rrect(1, 1, w-2, h-2, self._radius,
                        fill=self._fill, outline=self._outline, width=1)
            self._cv.tag_lower("bg")

        def set_colors(self, fill=None, outline=None):
            if fill is not None:
                self._fill = fill
                self.inner.configure(bg=fill)
            if outline is not None:
                self._outline = outline
            self._redraw()

    class RoundedButton(tk.Canvas):
        """Canvas button so main controls can have Win11-style rounded edges."""
        def __init__(self, parent, text, command=None, fill=None, hover=None, fg=None,
                     radius=12, padx=14, pady=7, font=("Segoe UI", 9, "bold"), **kw):
            self._bg = kw.pop("bg", None)
            if self._bg is None:
                try: self._bg = parent.cget("bg")
                except Exception: self._bg = "#000000"
            super().__init__(parent, bg=self._bg, highlightthickness=0, bd=0, **kw)
            self.text = text; self.command = command
            self.fill = fill or C["button"]; self.hover = hover or C["button_hover"]
            self.fg = fg or C["text"]; self.radius = radius
            self.padx = padx; self.pady = pady; self.font = font
            self.state = "normal"
            self._is_hover = False
            self.bind("<Configure>", lambda e: self._draw())
            self.bind("<Enter>", lambda e: self._set_hover(True))
            self.bind("<Leave>", lambda e: self._set_hover(False))
            self.bind("<Button-1>", self._click)
            self.after_idle(self._autosize)

        def _autosize(self):
            tmp = self.create_text(0, 0, text=self.text, font=self.font, anchor="nw")
            bb = self.bbox(tmp) or (0, 0, 80, 22)
            self.delete(tmp)
            # Call the real Tk Canvas configure here. Using self.config would
            # call this class override again and recurse forever on Python 3.13.
            tk.Canvas.configure(self, width=(bb[2]-bb[0]) + self.padx*2,
                                height=(bb[3]-bb[1]) + self.pady*2)
            self._draw()

        def _rrect(self, x1, y1, x2, y2, r, **kw):
            pts = [x1+r,y1, x2-r,y1, x2,y1, x2,y1+r, x2,y2-r, x2,y2,
                   x2-r,y2, x1+r,y2, x1,y2, x1,y2-r, x1,y1+r, x1,y1]
            self.create_polygon(pts, smooth=True, **kw)

        def _draw(self):
            self.delete("all")
            w, h = max(4, self.winfo_width()), max(4, self.winfo_height())
            col = self.hover if self._is_hover and self.state != "disabled" else self.fill
            if self.state == "disabled": col = C["card3"]
            self._rrect(1, 1, w-2, h-2, min(self.radius, h//2),
                        fill=col, outline=C["stroke_soft"], width=1)
            self.create_text(w//2, h//2, text=self.text, fill=(C["muted"] if self.state=="disabled" else self.fg),
                             font=self.font, anchor="center")

        def _set_hover(self, value):
            self._is_hover = value; self._draw()

        def _click(self, e):
            if self.state != "disabled" and self.command:
                self.command()

        def configure(self, cnf=None, **kw):
            if cnf is None:
                cnf = {}
            elif isinstance(cnf, str):
                # Preserve normal Tkinter behavior for widget.cget/configure("option") calls.
                return tk.Canvas.configure(self, cnf)
            else:
                cnf = dict(cnf)

            for source in (cnf, kw):
                if "text" in source: self.text = source.pop("text")
                if "command" in source: self.command = source.pop("command")
                if "state" in source: self.state = source.pop("state")

            result = tk.Canvas.configure(self, cnf, **kw)
            if cnf or kw:
                self._autosize()
            else:
                self._draw()
            return result
        config = configure

    try:
        from tkinterdnd2 import TkinterDnD, DND_FILES
        root = TkinterDnD.Tk(); HAS_DND = True
    except:
        root = tk.Tk(); HAS_DND = False

    cfg = load_config()

    # ── Windows 11 dark-blue glass / Mica palette ───────────────────────────
    # Tkinter cannot do true per-widget acrylic blur, so the theme uses a
    # dark translucent-glass palette plus native Win11 Mica/rounded corners
    # through DWM when available.
    C = {
        "bg":           "#0f1724",
        "sidebar":      "#0b1220",
        "card":         "#172235",
        "card2":        "#1b2a42",
        "card3":        "#132033",
        "stroke":       "#3b587a",
        "stroke_soft":  "#26364e",
        "text":         "#f6f8fb",
        "subtext":      "#d7e1ef",
        "muted":        "#a9bbd1",
        "accent":       "#60cdff",
        "green":        "#6cff6c",
        "warn":         "#ffd84d",
        "error":        "#ff7a7a",
        "field_bg":     "#101c2f",
        "field_fg":     "#f6f8fb",
        "field_border": "#3c526f",
        "button":       "#233a59",
        "button_hover": "#2f4e78",
        "sb_bg":        "#0b1220",
        "sb_thumb":     "#4f78a8",
    }

    # T() / ACC() shims — keep all original sidebar plumbing unchanged
    def T():
        return {"bg": C["bg"], "bg2": C["sidebar"], "bg3": C["button"],
                "fg": C["text"], "fg2": C["muted"], "sep": C["stroke_soft"]}
    def ACC(): return C["accent"]

    def card(parent, fill=None, outline=None, radius=16, pad=14, **kw):
        return RoundedCard(parent, fill=fill or C["card"],
                           outline=outline or C["stroke_soft"],
                           radius=radius, pad=pad, **kw)

    # diameter color map
    DIA_COLORS = {10:"#64b5f6",12:"#4fc3f7",16:"#81c784",20:"#e57373",
                  25:"#ffb74d",32:"#ce93d8",40:"#f06292"}

    def dia_color(d):
        try: return DIA_COLORS.get(int(d), ACC())
        except: return ACC()

    # weight/m lookup
    WPM = {6:0.222,8:0.395,10:0.617,12:0.888,16:1.578,20:2.466,25:3.854,32:6.313,40:9.865}
    def bar_weight(dia, total_len_mm, qty):
        try: return WPM.get(int(dia),0) * total_len_mm/1000 * qty
        except: return 0.0

    root.title("OPERA BBS Transformer by JK - v15")
    root.minsize(1100, 680)
    root.configure(bg=C["bg"])
    root.after(120, lambda: set_win11_glass(root))

    # ── app icon & logo ──
    _logo_photo = None  # keep reference alive
    try:
        from PIL import Image, ImageTk
        _logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png")
        if os.path.exists(_logo_path):
            _logo_img = Image.open(_logo_path).convert("RGBA")
            # Window icon (32x32)
            _icon = _logo_img.copy(); _icon.thumbnail((32,32), Image.LANCZOS)
            _icon_photo = ImageTk.PhotoImage(_icon)
            root.iconphoto(True, _icon_photo)
    except Exception:
        pass
    root.columnconfigure(0, weight=0, minsize=340)
    root.columnconfigure(1, weight=1)
    root.rowconfigure(0, weight=1)

    style = ttk.Style(root); style.theme_use("clam")

    def apply_theme():
        root.configure(bg=C["bg"])
        style.configure("TLabel",         background=C["bg"],      foreground=C["text"],   font=("Segoe UI",9))
        style.configure("TFrame",         background=C["bg"])
        style.configure("TSeparator",     background=C["stroke_soft"])
        style.configure("TLabelframe",    background=C["bg"],      foreground=C["accent"])
        style.configure("TLabelframe.Label", background=C["bg"],   foreground=C["accent"], font=("Segoe UI",9,"bold"))
        style.configure("TButton",        background=C["button"],  foreground=C["text"],   font=("Segoe UI",9,"bold"),
                        padding=(10,5),   borderwidth=1,            relief="flat",
                        bordercolor=C["stroke_soft"], lightcolor=C["button"], darkcolor=C["button"])
        style.map("TButton",              background=[("active",C["button_hover"]),("pressed","#3b5f91"),("disabled",C["card3"])],
                                          foreground=[("disabled",C["muted"])])
        style.configure("Accent.TButton", background=C["green"],   foreground="#071307",
                        font=("Segoe UI",9,"bold"), padding=(8,4), relief="flat",
                        bordercolor=C["green"], lightcolor=C["green"], darkcolor=C["green"])
        style.map("Accent.TButton",       background=[("active","#55ee55"),("pressed","#43d943"),("disabled",C["stroke"])])
        style.configure("TProgressbar",   troughcolor=C["card3"],  background=C["accent"],
                        bordercolor=C["stroke"], lightcolor=C["accent"], darkcolor=C["accent"], thickness=10)
        style.configure("TCombobox",      fieldbackground=C["field_bg"], background=C["field_bg"],
                        foreground=C["field_fg"], arrowcolor=C["accent"],
                        bordercolor=C["field_border"], selectbackground=C["button"],
                        selectforeground=C["text"], padding=4,
                        lightcolor=C["field_bg"], darkcolor=C["field_bg"])
        style.map("TCombobox",            fieldbackground=[("readonly",C["field_bg"])],
                                          foreground=[("readonly",C["field_fg"])],
                                          background=[("readonly",C["field_bg"])])
        try: root.update_idletasks()
        except: pass

    apply_theme()

    # ══════════════════════════════════════════════════════════════════════
    # LEFT SIDEBAR  (scrollable, minimal custom scrollbar)
    # ══════════════════════════════════════════════════════════════════════
    _sb_shell = tk.Frame(root, bg=C["sidebar"])
    _sb_shell.grid(row=0, column=0, sticky="nsew")
    _sb_shell.columnconfigure(0, weight=1)
    _sb_shell.rowconfigure(0, weight=1)

    _sb_canvas = tk.Canvas(_sb_shell, bg=C["sidebar"], highlightthickness=0, bd=0, width=340)
    _sb_canvas.grid(row=0, column=0, sticky="nsew")

    # Minimal custom scrollbar — 6 px wide, rounded thumb, no arrows
    _SB_W = 6
    _sb_bar = tk.Canvas(_sb_shell, width=_SB_W, bg=C["sidebar"],
                        highlightthickness=0, bd=0)
    _sb_bar.grid(row=0, column=1, sticky="ns")
    _sb_thumb_id = None

    def _sb_draw_thumb():
        nonlocal _sb_thumb_id
        _sb_bar.delete("all")
        try:
            lo, hi = _sb_canvas.yview()
        except Exception:
            return
        if hi - lo >= 1.0:          # nothing to scroll — hide thumb
            return
        h = _sb_bar.winfo_height() or 1
        ty = int(lo * h)
        th = max(24, int((hi - lo) * h))
        r  = _SB_W // 2
        col = C["sb_thumb"]
        # rounded rectangle via oval+rect
        _sb_bar.create_oval(0, ty,      _SB_W, ty + _SB_W,      fill=col, outline="")
        _sb_bar.create_oval(0, ty+th-_SB_W, _SB_W, ty+th, fill=col, outline="")
        _sb_bar.create_rectangle(0, ty+r, _SB_W, ty+th-r, fill=col, outline="")
        _sb_thumb_id = (ty, th)

    def _sb_set(lo, hi):
        _sb_canvas.yview_moveto(lo)
        _sb_bar.after_idle(_sb_draw_thumb)

    _sb_canvas.configure(yscrollcommand=_sb_set)
    _sb_bar.bind("<Configure>", lambda e: _sb_draw_thumb())

    # Drag the thumb
    _sb_drag = {}
    def _sb_bar_press(e):
        _sb_drag["y"] = e.y
        _sb_drag["lo"] = _sb_canvas.yview()[0]
    def _sb_bar_drag(e):
        h = _sb_bar.winfo_height() or 1
        delta = (e.y - _sb_drag["y"]) / h
        _sb_canvas.yview_moveto(_sb_drag["lo"] + delta)
        _sb_draw_thumb()
    _sb_bar.bind("<ButtonPress-1>",  _sb_bar_press)
    _sb_bar.bind("<B1-Motion>",      _sb_bar_drag)

    # Inner frame — all sidebar content lives here
    sidebar = tk.Frame(_sb_canvas, bg=C["sidebar"], padx=12, pady=14)
    _sb_win = _sb_canvas.create_window((0, 0), window=sidebar, anchor="nw")

    def _sb_on_configure(e):
        _sb_canvas.configure(scrollregion=_sb_canvas.bbox("all"))
        _sb_draw_thumb()
    def _sb_on_canvas_resize(e):
        _sb_canvas.itemconfig(_sb_win, width=e.width)
        _sb_draw_thumb()
    sidebar.bind("<Configure>", _sb_on_configure)
    _sb_canvas.bind("<Configure>", _sb_on_canvas_resize)

    def _sb_mousewheel(e):
        if e.num == 4 or (hasattr(e, 'delta') and e.delta > 0):
            _sb_canvas.yview_scroll(-1, "units")
        elif e.num == 5 or (hasattr(e, 'delta') and e.delta < 0):
            _sb_canvas.yview_scroll(1, "units")
        _sb_draw_thumb()
    for _w in (_sb_canvas, _sb_bar, sidebar):
        _w.bind("<MouseWheel>", _sb_mousewheel)
        _w.bind("<Button-4>",   _sb_mousewheel)
        _w.bind("<Button-5>",   _sb_mousewheel)

    sidebar.columnconfigure(0, weight=1)
    sr = 0  # sidebar row counter

    # Title row
    title_fr = tk.Frame(sidebar, bg=C["sidebar"])
    title_fr.grid(row=sr, column=0, sticky="ew"); sr+=1

    # Logo thumbnail in sidebar
    try:
        from PIL import Image, ImageTk
        _logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png")
        if os.path.exists(_logo_path):
            _thumb = Image.open(_logo_path).convert("RGBA")
            _thumb.thumbnail((48, 30), Image.LANCZOS)
            # Composite onto dark bg
            _bg = Image.new("RGBA", _thumb.size, (
                int(C["sidebar"][1:3],16),
                int(C["sidebar"][3:5],16),
                int(C["sidebar"][5:7],16), 255))
            _bg.paste(_thumb, mask=_thumb)
            _logo_photo = ImageTk.PhotoImage(_bg)
            tk.Label(title_fr, image=_logo_photo, bg=C["sidebar"], bd=0).pack(side="left", padx=(0,6))
            title_fr._logo_photo = _logo_photo  # keep reference
    except Exception:
        pass

    tk.Label(title_fr, text="OPERA BBS Transformer by JK", bg=C["sidebar"], fg=C["accent"],
             font=("Segoe UI",13,"bold")).pack(side="left")


    tk.Label(sidebar, text="Armabeton → Bar Bending Schedule.  Report bugs if found💩",
             bg=C["sidebar"], fg=C["muted"], font=("Segoe UI",8)).grid(
             row=sr, column=0, sticky="w", pady=(0,10)); sr+=1

    ttk.Separator(sidebar).grid(row=sr, column=0, sticky="ew", pady=(0,10)); sr+=1

    # ── recent files ──
    recent_var = tk.StringVar()
    recent_files = cfg.get("recent", [])

    def add_recent(path):
        if path in recent_files: recent_files.remove(path)
        recent_files.insert(0, path)
        del recent_files[5:]
        recent_combo["values"] = recent_files

    recent_fr = tk.Frame(sidebar, bg=C["sidebar"])
    recent_fr.grid(row=sr, column=0, sticky="ew", pady=(0,6)); sr+=1
    recent_fr.columnconfigure(0, weight=1)
    tk.Label(recent_fr, text="Recent Files", bg=C["sidebar"], fg=C["muted"],
             font=("Segoe UI",8)).grid(row=0, column=0, sticky="w")
    recent_combo = ttk.Combobox(recent_fr, textvariable=recent_var,
                                 values=recent_files, state="readonly", width=28)
    recent_combo.grid(row=1, column=0, sticky="ew", padx=(0,4))
    def on_recent(e):
        p = recent_var.get()
        if p and os.path.isfile(p): set_src(p)
    recent_combo.bind("<<ComboboxSelected>>", on_recent)
    RoundedButton(recent_fr, text="Open", command=lambda: on_recent(None),
                  fill=C["button"], hover=C["button_hover"], radius=10).grid(row=1, column=1, sticky="ns")

    ttk.Separator(sidebar).grid(row=sr, column=0, sticky="ew", pady=(6,8)); sr+=1

    # ── drop zone ──
    DROP_DEFAULT = "⬇  Drop .xls / .xlsx here\n     or click to browse"
    src_var = tk.StringVar(); dst_var = tk.StringVar()
    preview_excel_var = tk.BooleanVar(value=bool(cfg.get("excel_preview_sketches", False)))
    fit_sketches_to_cell_var = tk.BooleanVar(value=bool(cfg.get("fit_sketches_to_cell", False)))
    _drop_card = RoundedCard(sidebar, fill=C["card"], outline=C["stroke_soft"], radius=18, pad=0)
    _drop_card.grid(row=sr, column=0, sticky="ew", pady=(0,6)); sr+=1
    _drop_card.inner.columnconfigure(0, weight=1)
    drop_lbl = tk.Label(_drop_card.inner, text=DROP_DEFAULT, bg=C["card"], fg=C["accent"],
                        font=("Segoe UI",9), relief="flat", cursor="hand2",
                        justify="center", height=3)
    drop_lbl.grid(row=0, column=0, sticky="ew", padx=6, pady=6)

    def set_src(path):
        path = path.strip().strip("{}")
        if not os.path.isfile(path): return
        src_var.set(path)
        if not dst_var.get():
            dst_var.set(os.path.splitext(path)[0] + "_BBS.xlsx")
        drop_lbl.configure(text=f"✔  {os.path.basename(path)}",
                           bg=C["card"], fg=C["green"])
        _drop_card.set_colors(fill=C["card"], outline=C["green"])
        add_recent(path)
        threading.Thread(target=lambda: load_preview(path), daemon=True).start()

    if HAS_DND:
        drop_lbl.drop_target_register(DND_FILES)
        drop_lbl.dnd_bind("<<Drop>>",      lambda e: set_src(e.data))
        drop_lbl.dnd_bind("<<DragEnter>>", lambda e: drop_lbl.configure(bg=C["accent"], fg=C["bg"]))
        drop_lbl.dnd_bind("<<DragLeave>>", lambda e: drop_lbl.configure(
            bg=C["card"],
            fg=C["green"] if src_var.get() else C["accent"]))

    def browse_src():
        p = filedialog.askopenfilename(filetypes=[("Excel","*.xls *.xlsx"),("All","*.*")])
        if p: set_src(p)
    drop_lbl.bind("<Button-1>", lambda e: browse_src())

    out_fr = tk.Frame(sidebar, bg=C["sidebar"])
    out_fr.grid(row=sr, column=0, sticky="ew", pady=(0,10)); sr+=1
    out_fr.columnconfigure(0, weight=1)
    tk.Label(out_fr, text="Output .xlsx", bg=C["sidebar"], fg=C["muted"],
             font=("Segoe UI",8)).grid(row=0,column=0,sticky="w")
    tk.Entry(out_fr, textvariable=dst_var,
             bg=C["field_bg"], fg=C["field_fg"], insertbackground=C["accent"],
             relief="flat", bd=0, font=("Segoe UI",9), highlightthickness=1,
             highlightbackground=C["field_border"], highlightcolor=C["accent"]
             ).grid(row=1,column=0,sticky="ew",padx=(0,4))
    def browse_dst():
        p = filedialog.asksaveasfilename(defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")])
        if p: dst_var.set(p)
    RoundedButton(out_fr, text="…", command=browse_dst, fill=C["button"], hover=C["button_hover"], radius=10, padx=10).grid(row=1,column=1, sticky="ns")

    # ── Excel sketch source toggle ──
    sketch_opt = RoundedCard(sidebar, fill=C["card"], outline=C["stroke_soft"], radius=14, pad=8)
    sketch_opt.grid(row=sr, column=0, sticky="ew", pady=(0,8)); sr += 1
    sketch_opt.inner.columnconfigure(0, weight=1)
    tk.Checkbutton(sketch_opt.inner,
                   text="Use preview sketches in Excel",
                   variable=preview_excel_var,
                   bg=C["card"], fg=C["text"], selectcolor=C["field_bg"],
                   activebackground=C["card"], activeforeground=C["accent"],
                   font=("Segoe UI", 8, "bold"), relief="flat", bd=0,
                   highlightthickness=0).grid(row=0, column=0, sticky="w")
    tk.Label(sketch_opt.inner,
             text="On = generated sketches; Off = source WMF when available",
             bg=C["card"], fg=C["muted"], font=("Segoe UI", 7),
             anchor="w", wraplength=300, justify="left").grid(row=1, column=0, sticky="ew", pady=(2,0))
    tk.Checkbutton(sketch_opt.inner,
                   text="Fit sketches exactly to SKETCH cell",
                   variable=fit_sketches_to_cell_var,
                   bg=C["card"], fg=C["text"], selectcolor=C["field_bg"],
                   activebackground=C["card"], activeforeground=C["accent"],
                   font=("Segoe UI", 8, "bold"), relief="flat", bd=0,
                   highlightthickness=0).grid(row=2, column=0, sticky="w", pady=(6,0))
    tk.Label(sketch_opt.inner,
             text="Off = centered/aspect-ratio; On = full-cell exact fit",
             bg=C["card"], fg=C["muted"], font=("Segoe UI", 7),
             anchor="w", wraplength=300, justify="left").grid(row=3, column=0, sticky="ew", pady=(2,0))

    ttk.Separator(sidebar).grid(row=sr, column=0, sticky="ew", pady=(0,8)); sr+=1

    # ── header fields + presets ──
    hdr_hdr = tk.Frame(sidebar, bg=C["sidebar"])
    hdr_hdr.grid(row=sr, column=0, sticky="ew"); sr+=1
    hdr_hdr.columnconfigure(0, weight=1)
    hdr_open = tk.BooleanVar(value=True)
    def toggle_hdr():
        if hdr_open.get(): hf_inner.grid()
        else: hf_inner.grid_remove()
    hdr_toggle_btn = tk.Button(hdr_hdr, text="▼ Header Info", bg=C["sidebar"], fg=C["accent"],
                               font=("Segoe UI",9,"bold"), relief="flat", bd=0,
                               cursor="hand2", command=toggle_hdr)
    hdr_toggle_btn.grid(row=0, column=0, sticky="w")

    # Preset controls
    preset_var = tk.StringVar()
    presets = cfg.get("presets", {})
    preset_combo = ttk.Combobox(hdr_hdr, textvariable=preset_var,
                                 values=list(presets.keys()), width=12, state="readonly")
    preset_combo.grid(row=0, column=1, padx=4)

    hdr_vars = {}
    fields = [("Project:", "project_name"), ("BBS Name:", "bbs_name"),
              ("BBS No:", "bbs_no"),        ("SD No:", "sd_no"),
              ("Office:", "office"),         ("BBS Rev:", "bbs_rev"),
              ("SD Rev:", "sd_rev"),         ("Object:", "object_type"),
              ("Prep By:", "prepared_by"),   ("Chk By:", "checked_by")]

    # Header presets intentionally save/load only shared, non-unique fields.
    # Per-file unique fields stay editable in the batch popup:
    #   bbs_name, bbs_no, sd_no, bbs_rev, sd_rev
    HEADER_UNIQUE_KEYS = {"bbs_name", "bbs_no", "sd_no", "bbs_rev", "sd_rev"}
    HEADER_PRESET_KEYS = [key for _, key in fields if key not in HEADER_UNIQUE_KEYS]

    def load_preset(e=None):
        name = preset_var.get()
        if name in presets:
            saved = presets[name]
            # Support old presets, but only apply non-unique/shared fields.
            if isinstance(saved, dict) and "non_unique" in saved:
                saved = saved.get("non_unique", {})
            for k in HEADER_PRESET_KEYS:
                if k in hdr_vars and k in saved:
                    hdr_vars[k].set(saved.get(k, ""))

    def save_preset():
        name = preset_var.get().strip()
        if not name:
            name = filedialog.askstring("Preset Name", "Enter preset name:") if hasattr(filedialog, "askstring") else ""
            if not name:
                from tkinter.simpledialog import askstring
                name = askstring("Preset Name", "Enter name for this header preset:")
        if name:
            presets[name] = {
                "type": "non_unique_header",
                "non_unique": {k: hdr_vars[k].get() for k in HEADER_PRESET_KEYS if k in hdr_vars}
            }
            preset_combo["values"] = list(presets.keys())
            preset_var.set(name)

    preset_combo.bind("<<ComboboxSelected>>", load_preset)
    RoundedButton(hdr_hdr, text="Save Preset", command=save_preset, fill=C["button"], hover=C["button_hover"], radius=10, padx=10).grid(row=0, column=2)

    _hdr_card = RoundedCard(sidebar, fill=C["card"], outline=C["stroke_soft"], radius=18, pad=8)
    _hdr_card.grid(row=sr, column=0, sticky="ew", pady=(4,8)); sr+=1
    hf_inner = _hdr_card.inner
    hf_inner.columnconfigure(1, weight=1)
    for i, (lbl, key) in enumerate(fields):
        tk.Label(hf_inner, text=lbl, bg=C["card"], fg=C["muted"],
                 font=("Segoe UI",7), width=9, anchor="e").grid(
                 row=i, column=0, sticky="e", padx=(0,3), pady=1)
        v = tk.StringVar(); hdr_vars[key] = v
        tk.Entry(hf_inner, textvariable=v,
                 bg=C["field_bg"], fg=C["field_fg"], insertbackground=C["accent"],
                 relief="flat", bd=0, font=("Segoe UI",9), highlightthickness=1,
                 highlightbackground=C["field_border"], highlightcolor=C["accent"]
                 ).grid(row=i, column=1, sticky="ew", pady=1)

    ttk.Separator(sidebar).grid(row=sr, column=0, sticky="ew", pady=(0,8)); sr+=1

    # ── progress + buttons ──
    status_var = tk.StringVar(value="Ready.")
    tk.Label(sidebar, textvariable=status_var, bg=C["sidebar"], fg=C["muted"],
             font=("Segoe UI",8), wraplength=290, anchor="w").grid(
             row=sr, column=0, sticky="ew", pady=(0,4)); sr+=1
    progress = ttk.Progressbar(sidebar, maximum=100)
    progress.grid(row=sr, column=0, sticky="ew", pady=(0,10)); sr+=1

    btn_fr = tk.Frame(sidebar, bg=C["sidebar"])
    btn_fr.grid(row=sr, column=0, sticky="ew"); sr+=1
    btn_fr.columnconfigure(0, weight=1)
    convert_btn = RoundedButton(btn_fr, text="⚡  Convert", fill=C["green"], hover="#55ee55", fg="#071307", radius=14)
    convert_btn.grid(row=0, column=0, sticky="ew", pady=(0,4))
    def do_clear():
        src_var.set(""); dst_var.set("")
        [v.set("") for v in hdr_vars.values()]
        progress["value"]=0; status_var.set("Ready.")
        drop_lbl.configure(text=DROP_DEFAULT, bg=C["card"], fg=C["accent"])
        _drop_card.set_colors(fill=C["card"], outline=C["stroke_soft"])
        canvas.delete("all"); nav_label.configure(text="No file loaded")
        thumb_canvas.delete("all")
        for k in stat_vals: stat_vals[k].configure(text="—")
        ps.update({"bars":[],"idx":0,"wmf_map":{},"photos":[]})
        # also clear batch queue
        batch_queue.clear()
        queue_lb.delete(0, "end")
        batch_log_var.set("")
        batch_progress["value"] = 0
    RoundedButton(btn_fr, text="Clear", command=do_clear, fill=C["button"], hover=C["button_hover"], radius=14).grid(row=1,column=0,sticky="ew")

    # ── batch toggle button ──
    ttk.Separator(sidebar).grid(row=sr, column=0, sticky="ew", pady=(8,4)); sr+=1
    _batch_toggle_row = sr - 1  # row of the separator/toggle button
    batch_open = tk.BooleanVar(value=False)
    _batch_card = RoundedCard(sidebar, fill=C["card"], outline=C["stroke_soft"], radius=18, pad=8)
    _batch_card.inner.columnconfigure(0, weight=1)
    batch_section = _batch_card.inner  # alias so existing code unchanged
    _batch_row = sr  # batch section goes at current sr

    def toggle_batch(_row=_batch_row):
        if batch_open.get():
            _batch_card.grid(row=_row, column=0, sticky="ew", pady=(0,6))
        else:
            _batch_card.grid_remove()

    tk.Button(sidebar, text="⚡ Batch Convert ▼", bg=C["sidebar"], fg=C["accent"],
              font=("Segoe UI",9,"bold"), relief="flat", bd=0, cursor="hand2",
              command=lambda: (batch_open.set(not batch_open.get()), toggle_batch())
              ).grid(row=_batch_toggle_row, column=0, sticky="w")

    # ── batch queue list ──
    queue_lbl = tk.Label(batch_section, text="Queue (source files):",
                         bg=C["card"], fg=C["muted"], font=("Segoe UI",8))
    queue_lbl.grid(row=0, column=0, sticky="w")

    queue_frame = tk.Frame(batch_section, bg=C["card"])
    queue_frame.grid(row=1, column=0, sticky="nsew")
    queue_frame.columnconfigure(0, weight=1)
    batch_section.rowconfigure(1, weight=0)

    queue_lb = tk.Listbox(queue_frame, bg=C["card2"], fg=C["text"],
                          selectbackground=C["button_hover"], font=("Segoe UI",8),
                          activestyle="none", relief="flat", bd=0, highlightthickness=0, height=6)
    queue_lb.pack(side="left", fill="both", expand=True)
    # Minimal canvas scrollbar — no arrows, themed thumb
    _QL_W = 5
    _ql_sb = tk.Canvas(queue_frame, width=_QL_W, bg=C["card2"], highlightthickness=0, bd=0)
    _ql_sb.pack(side="right", fill="y")
    def _ql_draw(lo=None, hi=None):
        _ql_sb.delete("all")
        try: lo, hi = (float(lo), float(hi)) if lo is not None else queue_lb.yview()
        except: return
        if hi - lo >= 1.0: return
        h = _ql_sb.winfo_height() or 1
        ty = int(lo * h); th = max(16, int((hi - lo) * h))
        r = _QL_W // 2; col = C["sb_thumb"]
        _ql_sb.create_oval(0,ty,_QL_W,ty+_QL_W,fill=col,outline="")
        _ql_sb.create_oval(0,ty+th-_QL_W,_QL_W,ty+th,fill=col,outline="")
        _ql_sb.create_rectangle(0,ty+r,_QL_W,ty+th-r,fill=col,outline="")
    queue_lb.configure(yscrollcommand=_ql_draw)
    _ql_sb.bind("<Configure>", lambda e: _ql_draw())
    _ql_drag = {}
    def _ql_press(e): _ql_drag["y"]=e.y; _ql_drag["lo"]=queue_lb.yview()[0]
    def _ql_move(e):
        h=_ql_sb.winfo_height() or 1; d=(e.y-_ql_drag["y"])/h
        queue_lb.yview_moveto(_ql_drag["lo"]+d); _ql_draw()
    _ql_sb.bind("<ButtonPress-1>",_ql_press); _ql_sb.bind("<B1-Motion>",_ql_move)

    # Batch queue items keep optional per-file header overrides.
    # Only these five fields are unique per queued file. Everything else
    # (project, office, prepared/checked by, object type, etc.) is inherited
    # from the main Header Info section at run time. Batch filenames use
    # BBS No as the output filename plus the BBS Rev suffix, defaulting to -00.
    BATCH_UNIQUE_KEYS = ["bbs_name", "bbs_no", "sd_no", "bbs_rev", "sd_rev"]
    BATCH_UNIQUE_LABELS = [
        ("BBS Name", "bbs_name"),
        ("BBS No", "bbs_no"),
        ("SD No", "sd_no"),
        ("BBS Rev", "bbs_rev"),
        ("Shop Drawing Rev", "sd_rev"),
    ]
    batch_queue = []  # list of {"src": path, "hdr": {only the five unique overrides}}
    batch_selected_idx = {"idx": -1}

    def clean_unique_hdr(hdr):
        """Keep only the allowed per-file unique fields."""
        hdr = hdr or {}
        return {k: str(hdr.get(k, "")).strip() for k in BATCH_UNIQUE_KEYS if str(hdr.get(k, "")).strip()}

    def safe_filename_part(value):
        """Return a Windows-safe filename component."""
        value = str(value or "").strip()
        value = re.sub(r'[<>:"/\\|?*]+', '-', value)
        value = re.sub(r'\s+', ' ', value).strip(' .-_')
        return value

    def normalize_revision(value):
        """Return a safe BBS revision suffix. Blank revision defaults to 00."""
        rev = safe_filename_part(value)
        if not rev:
            return "00"
        rev = rev.strip()
        if rev.upper().startswith("REV"):
            rev = rev[3:].strip(" .-_") or "00"
        return rev

    def append_revision_to_filename(base_name, bbs_rev):
        """Use BBS No as the filename base and append the BBS revision.

        Example: BV-BSW-105-0000-ECM-BBS-001 + 00 ->
        BV-BSW-105-0000-ECM-BBS-001-00. If the base already ends with the
        same revision suffix, it is not duplicated.
        """
        base = safe_filename_part(base_name)
        rev = normalize_revision(bbs_rev)
        if not base:
            return ""
        if base.upper().endswith(("-" + rev).upper()):
            return base
        return f"{base}-{rev}"

    def output_stem_for_item(item, header_defaults=None):
        """Batch output name rule: BBS No is the file name, plus BBS Rev.

        The popup stores BBS Name, BBS No, SD No, BBS Rev, and Shop Drawing Rev
        as per-file unique data. If a queued file has no BBS Rev, the main Header
        Info BBS Rev is used; if still blank, the filename suffix defaults to -00.
        """
        header_defaults = header_defaults or {}
        hdr = clean_unique_hdr(item.get("hdr", {}))
        rev = hdr.get("bbs_rev") or header_defaults.get("bbs_rev")
        stem = append_revision_to_filename(hdr.get("bbs_no"), rev)
        if stem:
            return stem
        # Fallback if BBS No is blank: keep the source file stem and add the rev.
        src_stem = safe_filename_part(os.path.splitext(os.path.basename(item.get("src", "")))[0])
        rev = normalize_revision(rev)
        if src_stem and not src_stem.upper().endswith(("-" + rev).upper()):
            return f"{src_stem}-{rev}"
        return src_stem or f"BBS-{rev}"

    def unique_output_path(folder, stem):
        """Avoid overwriting previous batch outputs with the same BBS Name/revision."""
        stem = safe_filename_part(stem) or "BBS-00"
        candidate = os.path.join(folder, stem + ".xlsx")
        if not os.path.exists(candidate):
            return candidate
        n = 2
        while True:
            candidate = os.path.join(folder, f"{stem} ({n}).xlsx")
            if not os.path.exists(candidate):
                return candidate
            n += 1

    def batch_item_label(item):
        hdr = clean_unique_hdr(item.get("hdr", {}))
        base = os.path.basename(item.get("src", ""))
        bits = []
        if hdr.get("bbs_name"):
            bits.append(f"BBS Name: {hdr['bbs_name']}")
        if hdr.get("bbs_no"):
            bits.append(f"BBS No: {hdr['bbs_no']}")
        if hdr.get("sd_no"):
            bits.append(f"SD No: {hdr['sd_no']}")
        if hdr.get("bbs_rev"):
            bits.append(f"BBS Rev: {hdr['bbs_rev']}")
        if hdr.get("sd_rev"):
            bits.append(f"SD Rev: {hdr['sd_rev']}")
        return base if not bits else base + "  |  " + "  |  ".join(bits)

    def refresh_batch_list(keep_selection=True):
        cur = batch_selected_idx["idx"] if keep_selection else -1
        queue_lb.delete(0, "end")
        for item in batch_queue:
            queue_lb.insert("end", batch_item_label(item))
        if keep_selection and 0 <= cur < len(batch_queue):
            queue_lb.selection_set(cur)
            queue_lb.see(cur)

    def load_batch_selection(idx):
        batch_selected_idx["idx"] = idx
        if 0 <= idx < len(batch_queue):
            batch_log_var.set(os.path.basename(batch_queue[idx].get("src", "")) + " selected. Click Edit Info to rename.")
        else:
            batch_log_var.set("")

    def on_batch_select(e=None):
        sel = list(queue_lb.curselection())
        load_batch_selection(sel[0] if sel else -1)

    def default_batch_hdr_for_path(path):
        # Use the source filename as an initial BBS No because batch output filenames
        # are based on BBS No + BBS Rev. The user can edit it in the popup.
        stem = safe_filename_part(os.path.splitext(os.path.basename(path))[0])
        return {"bbs_no": stem, "bbs_rev": "00"} if stem else {"bbs_rev": "00"}

    def apply_batch_popup_values(rows, popup=None):
        for row in rows:
            item = row["item"]
            item["hdr"] = {k: row["vars"][k].get().strip() for k in BATCH_UNIQUE_KEYS if row["vars"][k].get().strip()}
        refresh_batch_list(keep_selection=True)
        if popup is not None:
            popup.destroy()

    def batch_bbs_name_from_file(rows):
        # Direction: source file stem -> BBS Name field.
        for row in rows:
            stem = safe_filename_part(os.path.splitext(os.path.basename(row["item"].get("src", "")))[0])
            row["vars"]["bbs_name"].set(stem)

    def batch_bbs_no_from_file(rows):
        # Direction: source file stem -> BBS No field.
        for row in rows:
            stem = safe_filename_part(os.path.splitext(os.path.basename(row["item"].get("src", "")))[0])
            row["vars"]["bbs_no"].set(stem)


    def batch_file_to_number_fields(rows):
        # Direction: source file stem -> BBS No and SD No.
        # The filename preview follows automatically because batch filenames use BBS No + BBS Rev.
        for row in rows:
            stem = safe_filename_part(os.path.splitext(os.path.basename(row["item"].get("src", "")))[0])
            row["vars"]["bbs_no"].set(stem)
            row["vars"]["sd_no"].set(stem)

    def batch_sd_no_from_bbs_no(rows):
        # Direction: BBS No field -> SD No field.
        for row in rows:
            row["vars"]["sd_no"].set(row["vars"].get("bbs_no", tk.StringVar(value="")).get().strip())

    def batch_sd_rev_from_bbs_rev(rows):
        # Direction: BBS Rev field -> Shop Drawing Rev field.
        for row in rows:
            row["vars"]["sd_rev"].set(row["vars"].get("bbs_rev", tk.StringVar(value="")).get().strip())

    def batch_sd_fields_from_bbs_fields(rows):
        # Combined helper: BBS No -> SD No and BBS Rev -> Shop Drawing Rev.
        batch_sd_no_from_bbs_no(rows)
        batch_sd_rev_from_bbs_rev(rows)

    def clear_batch_rows(rows):
        for row in rows:
            for key in BATCH_UNIQUE_KEYS:
                row["vars"][key].set("")

    def open_batch_info_popup(indices=None):
        if not batch_queue:
            messagebox.showwarning("Batch Info", "Queue is empty - add files first.")
            return
        if indices is None:
            indices = list(range(len(batch_queue)))
        indices = [i for i in indices if 0 <= i < len(batch_queue)]
        if not indices:
            messagebox.showwarning("Batch Info", "Select a queued file first.")
            return

        popup = tk.Toplevel(root)
        popup.title("Batch BBS Info")
        popup.configure(bg=C["bg"])
        popup.geometry("960x520")
        popup.minsize(760, 360)
        popup.transient(root)
        popup.grab_set()
        popup.columnconfigure(0, weight=1)
        popup.rowconfigure(1, weight=1)
        popup.after(120, lambda: set_win11_glass(popup))

        tk.Label(
            popup,
            text="Batch rename table: BBS Name uses Drawing Title, BBS No / SD No uses Detected Drawing No, and BBS Rev / Shop Drawing Rev uses Append. JSON export uses Drawing = null_bbs_source and Layout = original XLS/XLSX file name. Output filename uses BBS No + BBS Rev, for example BBS-NO-00.xlsx.",
            bg=C["bg"], fg=C["subtext"], font=("Segoe UI", 9), anchor="w", justify="left",
            wraplength=900
        ).grid(row=0, column=0, sticky="ew", padx=14, pady=(12, 6))

        body = tk.Frame(popup, bg=C["card"])
        body.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 8))
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        canvas2 = tk.Canvas(body, bg=C["card"], highlightthickness=0, bd=0)
        canvas2.grid(row=0, column=0, sticky="nsew")

        # Themed popup scrollbar - same slim canvas style used by the sidebar
        # and queue list.  This avoids the default ttk scrollbar looking out of
        # place against the dark Win11-style theme.
        POP_SB_W = 6
        popup_scrollbar = tk.Canvas(body, width=POP_SB_W, bg=C["card"],
                                    highlightthickness=0, bd=0)
        popup_scrollbar.grid(row=0, column=1, sticky="ns")
        popup_scroll_state = {"drag_y": 0, "drag_lo": 0.0}

        grid = tk.Frame(canvas2, bg=C["card"])
        win_id = canvas2.create_window((0, 0), window=grid, anchor="nw")

        def draw_popup_scrollbar():
            popup_scrollbar.delete("all")
            try:
                lo, hi = canvas2.yview()
            except Exception:
                return
            if hi - lo >= 1.0:
                return
            h = popup_scrollbar.winfo_height() or 1
            thumb_y = int(lo * h)
            thumb_h = max(24, int((hi - lo) * h))
            radius = POP_SB_W // 2
            color = C["sb_thumb"]
            popup_scrollbar.create_oval(0, thumb_y, POP_SB_W, thumb_y + POP_SB_W,
                                        fill=color, outline="")
            popup_scrollbar.create_oval(0, thumb_y + thumb_h - POP_SB_W, POP_SB_W, thumb_y + thumb_h,
                                        fill=color, outline="")
            popup_scrollbar.create_rectangle(0, thumb_y + radius, POP_SB_W, thumb_y + thumb_h - radius,
                                             fill=color, outline="")

        def popup_yview(*args):
            canvas2.yview(*args)
            draw_popup_scrollbar()

        def on_popup_scrollbar_press(e):
            popup_scroll_state["drag_y"] = e.y
            popup_scroll_state["drag_lo"] = canvas2.yview()[0]

        def on_popup_scrollbar_drag(e):
            h = popup_scrollbar.winfo_height() or 1
            delta = (e.y - popup_scroll_state["drag_y"]) / h
            canvas2.yview_moveto(max(0.0, min(1.0, popup_scroll_state["drag_lo"] + delta)))
            draw_popup_scrollbar()

        def on_popup_mousewheel(e):
            if e.num == 4 or (hasattr(e, "delta") and e.delta > 0):
                canvas2.yview_scroll(-1, "units")
            elif e.num == 5 or (hasattr(e, "delta") and e.delta < 0):
                canvas2.yview_scroll(1, "units")
            draw_popup_scrollbar()

        popup_scrollbar.bind("<Configure>", lambda e: draw_popup_scrollbar())
        popup_scrollbar.bind("<ButtonPress-1>", on_popup_scrollbar_press)
        popup_scrollbar.bind("<B1-Motion>", on_popup_scrollbar_drag)
        popup_scrollbar.bind("<MouseWheel>", on_popup_mousewheel)
        popup_scrollbar.bind("<Button-4>", on_popup_mousewheel)
        popup_scrollbar.bind("<Button-5>", on_popup_mousewheel)
        canvas2.configure(yscrollcommand=lambda lo, hi: draw_popup_scrollbar())

        def on_grid_configure(e=None):
            canvas2.configure(scrollregion=canvas2.bbox("all"))
            draw_popup_scrollbar()
        def on_canvas_configure(e):
            canvas2.itemconfigure(win_id, width=e.width)
            draw_popup_scrollbar()
        grid.bind("<Configure>", on_grid_configure)
        canvas2.bind("<Configure>", on_canvas_configure)
        for _w in (popup, body, canvas2, grid):
            _w.bind("<MouseWheel>", on_popup_mousewheel)
            _w.bind("<Button-4>", on_popup_mousewheel)
            _w.bind("<Button-5>", on_popup_mousewheel)

        headers = ["File", "BBS Name / Drawing Title", "BBS No / SD No / Filename", "BBS Rev / Shop Drawing Rev / Append"]
        widths = [30, 30, 34, 26]
        row_line_color = "#3a4d66"  # subtle light divider, consistent with the dark theme
        for ci, title in enumerate(headers):
            tk.Label(grid, text=title, bg=C["card2"], fg=C["accent"], font=("Segoe UI", 8, "bold"),
                     anchor="w", padx=6, pady=5).grid(row=0, column=ci, sticky="ew", padx=0, pady=0)
            grid.columnconfigure(ci, weight=1)
        # Only horizontal separators are drawn.  Columns stay visually clean,
        # while each batch row remains easy to distinguish in the dark popup.
        tk.Frame(grid, bg=row_line_color, height=1).grid(row=1, column=0, columnspan=len(headers), sticky="ew")

        rows = []
        for ri, idx in enumerate(indices, 1):
            grid_row = ri * 2
            separator_row = grid_row + 1
            item = batch_queue[idx]
            hdr = clean_unique_hdr(item.get("hdr", {}))
            vars_for_row = {k: tk.StringVar(value=hdr.get(k, "")) for k in BATCH_UNIQUE_KEYS}
            preview_var = tk.StringVar()

            row_rec = {"item": item, "vars": vars_for_row, "preview": preview_var}
            rows.append(row_rec)

            tk.Label(grid, text=os.path.basename(item.get("src", "")), bg=C["card"], fg=C["text"],
                     font=("Segoe UI", 8), anchor="w", padx=6, wraplength=220).grid(row=grid_row, column=0, sticky="nsew", padx=0, pady=(4, 4))

            ent_name = tk.Entry(grid, textvariable=vars_for_row["bbs_name"], width=widths[1],
                                bg=C["field_bg"], fg=C["field_fg"], insertbackground=C["accent"],
                                relief="flat", bd=0, font=("Segoe UI", 8), highlightthickness=1,
                                highlightbackground=C["field_border"], highlightcolor=C["accent"])
            ent_name.grid(row=grid_row, column=1, sticky="ew", padx=(6, 6), pady=(4, 4))

            number_frame = tk.Frame(grid, bg=C["card"])
            number_frame.grid(row=grid_row, column=2, sticky="nsew", padx=(6, 6), pady=(4, 4))
            number_frame.columnconfigure(1, weight=1)
            tk.Label(number_frame, text="BBS No", bg=C["card"], fg=C["muted"], font=("Segoe UI", 7), width=8, anchor="e").grid(row=0, column=0, sticky="e", padx=(0, 4), pady=1)
            tk.Entry(number_frame, textvariable=vars_for_row["bbs_no"], bg=C["field_bg"], fg=C["field_fg"], insertbackground=C["accent"],
                     relief="flat", bd=0, font=("Segoe UI", 8), highlightthickness=1,
                     highlightbackground=C["field_border"], highlightcolor=C["accent"]).grid(row=0, column=1, sticky="ew", pady=1)
            tk.Label(number_frame, text="SD No", bg=C["card"], fg=C["muted"], font=("Segoe UI", 7), width=8, anchor="e").grid(row=1, column=0, sticky="e", padx=(0, 4), pady=1)
            tk.Entry(number_frame, textvariable=vars_for_row["sd_no"], bg=C["field_bg"], fg=C["field_fg"], insertbackground=C["accent"],
                     relief="flat", bd=0, font=("Segoe UI", 8), highlightthickness=1,
                     highlightbackground=C["field_border"], highlightcolor=C["accent"]).grid(row=1, column=1, sticky="ew", pady=1)
            tk.Label(number_frame, text="Filename", bg=C["card"], fg=C["muted"], font=("Segoe UI", 7), width=8, anchor="e").grid(row=2, column=0, sticky="e", padx=(0, 4), pady=1)
            tk.Label(number_frame, textvariable=preview_var, bg=C["card"], fg=C["accent"],
                     font=("Segoe UI", 8), anchor="w").grid(row=2, column=1, sticky="ew", pady=1)

            rev_frame = tk.Frame(grid, bg=C["card"])
            rev_frame.grid(row=grid_row, column=3, sticky="nsew", padx=(6, 6), pady=(4, 4))
            rev_frame.columnconfigure(1, weight=1)
            tk.Label(rev_frame, text="BBS Rev", bg=C["card"], fg=C["muted"], font=("Segoe UI", 7), width=12, anchor="e").grid(row=0, column=0, sticky="e", padx=(0, 4), pady=1)
            tk.Entry(rev_frame, textvariable=vars_for_row["bbs_rev"], bg=C["field_bg"], fg=C["field_fg"], insertbackground=C["accent"],
                     relief="flat", bd=0, font=("Segoe UI", 8), highlightthickness=1,
                     highlightbackground=C["field_border"], highlightcolor=C["accent"]).grid(row=0, column=1, sticky="ew", pady=1)
            tk.Label(rev_frame, text="Shop DWG Rev", bg=C["card"], fg=C["muted"], font=("Segoe UI", 7), width=12, anchor="e").grid(row=1, column=0, sticky="e", padx=(0, 4), pady=1)
            tk.Entry(rev_frame, textvariable=vars_for_row["sd_rev"], bg=C["field_bg"], fg=C["field_fg"], insertbackground=C["accent"],
                     relief="flat", bd=0, font=("Segoe UI", 8), highlightthickness=1,
                     highlightbackground=C["field_border"], highlightcolor=C["accent"]).grid(row=1, column=1, sticky="ew", pady=1)

            tk.Frame(grid, bg=row_line_color, height=1).grid(
                row=separator_row, column=0, columnspan=len(headers), sticky="ew"
            )

            def update_preview(*args, _row=row_rec):
                fake = {"src": _row["item"].get("src", ""), "hdr": {k: v.get() for k, v in _row["vars"].items()}}
                _row["preview"].set(output_stem_for_item(fake, {"bbs_rev": hdr_vars.get("bbs_rev", tk.StringVar(value="00")).get()}) + ".xlsx")
            for v in vars_for_row.values():
                v.trace_add("write", update_preview)
            update_preview()

        def import_rename_json():
            path = filedialog.askopenfilename(
                title="Import BBS rename JSON",
                filetypes=[("JSON", "*.json"), ("All", "*.*")])
            if not path:
                return
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                src_rows = data.get("rows", []) if isinstance(data, dict) else []
                if not isinstance(src_rows, list) or not src_rows:
                    raise ValueError("No rows found in JSON file.")

                def _norm_key(value):
                    value = str(value or "").strip()
                    value = os.path.splitext(os.path.basename(value))[0]
                    return safe_filename_part(value).lower()

                def _apply_import(target_row, src):
                    drawing_title = str(src.get("Drawing Title") or src.get("BBS Name") or "").strip()
                    detected_no = str(src.get("Detected Drawing No") or src.get("BBS No") or src.get("SD No") or "").strip()
                    append = str(src.get("Append") or src.get("BBS Rev") or src.get("Shop Drawing Rev") or src.get("SD Rev") or "").strip()

                    if drawing_title:
                        target_row["vars"]["bbs_name"].set(drawing_title)
                    if detected_no:
                        target_row["vars"]["bbs_no"].set(detected_no)
                        target_row["vars"]["sd_no"].set(detected_no)
                    if append:
                        target_row["vars"]["bbs_rev"].set(append)
                        target_row["vars"]["sd_rev"].set(append)

                fmt = str(data.get("format", "") if isinstance(data, dict) else "").strip()
                matched = 0
                used_json_rows = set()

                # CAD Plot Center PDF Rename imports are matched by Layout only.
                # This lets the popup fill the correct queued file even when the
                # JSON order differs from the batch order. The Layout text is
                # compared against the original XLS/XLSX filename stem.
                if fmt.lower() == "cad plot center pdf rename":
                    json_by_layout = {}
                    for src in src_rows:
                        if not isinstance(src, dict):
                            continue
                        layout_key = _norm_key(src.get("Layout", ""))
                        if layout_key and layout_key not in json_by_layout:
                            json_by_layout[layout_key] = src

                    for row in rows:
                        source_key = _norm_key(row["item"].get("src", ""))
                        src = json_by_layout.get(source_key)
                        if src:
                            _apply_import(row, src)
                            matched += 1
                            used_json_rows.add(id(src))

                    messagebox.showinfo(
                        "Import Rename JSON",
                        f"Imported {matched} row(s) by Layout.\nUnmatched popup row(s): {len(rows) - matched}."
                    )
                    return

                # BBS Transformer Rename imports also prefer Layout-to-filename
                # matching when possible, then fall back to visible row order.
                json_by_layout = {}
                for src in src_rows:
                    if not isinstance(src, dict):
                        continue
                    layout_key = _norm_key(src.get("Layout", ""))
                    if layout_key and layout_key not in json_by_layout:
                        json_by_layout[layout_key] = src

                for row in rows:
                    source_key = _norm_key(row["item"].get("src", ""))
                    src = json_by_layout.get(source_key)
                    if src:
                        _apply_import(row, src)
                        matched += 1
                        used_json_rows.add(id(src))

                if matched == 0:
                    count = min(len(rows), len(src_rows))
                    for i in range(count):
                        src = src_rows[i] if isinstance(src_rows[i], dict) else {}
                        _apply_import(rows[i], src)
                    matched = count
                    messagebox.showinfo("Import Rename JSON", f"Imported {matched} row(s) by row order.")
                else:
                    messagebox.showinfo("Import Rename JSON", f"Imported {matched} row(s) by Layout.")
            except Exception as e:
                messagebox.showerror("Import Rename JSON", str(e))

        def export_rename_json():
            path = filedialog.asksaveasfilename(
                title="Export BBS rename JSON",
                defaultextension=".json",
                filetypes=[("JSON", "*.json")])
            if not path:
                return
            try:
                out_rows = []
                for i, row in enumerate(rows, 1):
                    v = row["vars"]
                    bbs_name = v["bbs_name"].get().strip()
                    bbs_no = v["bbs_no"].get().strip()
                    sd_no = v["sd_no"].get().strip()
                    bbs_rev = normalize_revision(v["bbs_rev"].get().strip())
                    sd_rev = v["sd_rev"].get().strip() or bbs_rev
                    detected_no = sd_no or bbs_no
                    append = sd_rev or bbs_rev
                    pdf_name = append_revision_to_filename(detected_no, append) if detected_no else ""
                    source_name = safe_filename_part(os.path.splitext(os.path.basename(row["item"].get("src", "")))[0])
                    out_rows.append({
                        "No": i,
                        # Keep this compatible with the rename software format:
                        # Drawing is a fixed source tag, while Layout carries the original XLS/XLSX file name.
                        "Drawing": "null_bbs_source",
                        "Layout": source_name,
                        "Detected Drawing No": detected_no,
                        "Append": append,
                        "PDF Filename": pdf_name,
                        "Drawing Title": bbs_name,
                    })

                payload = {
                    "format": "BBS Transformer Rename",
                    "version": 1,
                    "rows": out_rows,
                }
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(payload, f, indent=2, ensure_ascii=False)
                messagebox.showinfo("Export Rename JSON", f"Exported {len(out_rows)} row(s).")
            except Exception as e:
                messagebox.showerror("Export Rename JSON", str(e))

        btns = tk.Frame(popup, bg=C["bg"])
        btns.grid(row=2, column=0, sticky="ew", padx=14, pady=(0, 12))
        btns.columnconfigure(0, weight=1)
        RoundedButton(btns, text="Import JSON", command=import_rename_json, radius=10, padx=12, pady=6).pack(side="left", padx=(0, 6))
        RoundedButton(btns, text="Export JSON", command=export_rename_json, radius=10, padx=12, pady=6).pack(side="left", padx=(0, 6))
        RoundedButton(btns, text="File -> BBS Name", command=lambda: batch_bbs_name_from_file(rows), radius=10, padx=12, pady=6).pack(side="left", padx=(0, 6))
        RoundedButton(btns, text="BBS No/Rev -> SD No/Shop DWG Rev", command=lambda: batch_sd_fields_from_bbs_fields(rows), radius=10, padx=12, pady=6).pack(side="left", padx=(0, 6))
        RoundedButton(btns, text="Save", command=lambda: apply_batch_popup_values(rows, popup), fill=C["green"], hover="#55ee55", fg="#071307", radius=10, padx=18, pady=6).pack(side="right", padx=(6, 0))
        RoundedButton(btns, text="Cancel", command=popup.destroy, radius=10, padx=16, pady=6).pack(side="right")

        popup.bind("<Escape>", lambda e: popup.destroy())
        popup.bind("<Control-s>", lambda e: apply_batch_popup_values(rows, popup))
        popup.after(50, lambda: popup.focus_force())

    def batch_add_files():
        paths = filedialog.askopenfilenames(
            title="Add files to batch",
            filetypes=[("Excel","*.xls *.xlsx"),("All","*.*")])
        existing = {item["src"] for item in batch_queue}
        new_indices = []
        for p in paths:
            if p not in existing:
                batch_queue.append({"src": p, "hdr": default_batch_hdr_for_path(p)})
                existing.add(p)
                new_indices.append(len(batch_queue) - 1)
        if paths:
            if batch_selected_idx["idx"] < 0 and batch_queue:
                batch_selected_idx["idx"] = 0
            refresh_batch_list(keep_selection=True)
            if new_indices:
                batch_selected_idx["idx"] = new_indices[0]
                refresh_batch_list(keep_selection=True)
                open_batch_info_popup(new_indices)

    def batch_edit_selected():
        sel = list(queue_lb.curselection())
        if not sel:
            messagebox.showwarning("Batch Info", "Select one or more queued files to edit.")
            return
        open_batch_info_popup(sel)

    def batch_edit_all():
        open_batch_info_popup(list(range(len(batch_queue))))

    def batch_remove_sel():
        sel = list(queue_lb.curselection())
        for i in reversed(sel):
            del batch_queue[i]
        new_idx = min(sel[0], len(batch_queue)-1) if sel else -1
        batch_selected_idx["idx"] = new_idx
        refresh_batch_list(keep_selection=True)
        load_batch_selection(new_idx)

    def batch_clear_queue():
        batch_queue.clear()
        queue_lb.delete(0, "end")
        batch_selected_idx["idx"] = -1
        load_batch_selection(-1)

    queue_lb.bind("<<ListboxSelect>>", on_batch_select)
    queue_lb.bind("<Double-Button-1>", lambda e: batch_edit_selected())

    qbtn_fr = tk.Frame(batch_section, bg=C["card"])
    qbtn_fr.grid(row=2, column=0, sticky="ew", pady=(2,4))
    RoundedButton(qbtn_fr, text="+ Add", command=batch_add_files, radius=10, padx=10, pady=5).pack(side="left", padx=(0,4))
    RoundedButton(qbtn_fr, text="Edit Info", command=batch_edit_selected, radius=10, padx=10, pady=5).pack(side="left", padx=(0,4))
    RoundedButton(qbtn_fr, text="Edit All", command=batch_edit_all, radius=10, padx=10, pady=5).pack(side="left", padx=(0,4))
    RoundedButton(qbtn_fr, text="Remove", command=batch_remove_sel, radius=10, padx=10, pady=5).pack(side="left", padx=(0,4))
    RoundedButton(qbtn_fr, text="Clear", command=batch_clear_queue, radius=10, padx=10, pady=5).pack(side="left")

    tk.Label(batch_section,
             text="Double-click a queued file or use Edit Info. Batch output filenames use BBS No + BBS Rev, for example BBS-NO-00.xlsx.",
             bg=C["card"], fg=C["muted"], font=("Segoe UI", 7), anchor="w", wraplength=300,
             justify="left").grid(row=3, column=0, sticky="ew", pady=(0,6))

    # Output folder for batch
    batch_out_fr = tk.Frame(batch_section, bg=C["card"])
    batch_out_fr.grid(row=4, column=0, sticky="ew", pady=(0,4))
    batch_out_fr.columnconfigure(1, weight=1)
    tk.Label(batch_out_fr, text="Out folder:", bg=C["card"], fg=C["muted"],
             font=("Segoe UI",8)).grid(row=0, column=0, sticky="w", padx=(0,4))
    batch_outdir_var = tk.StringVar(value="")
    tk.Entry(batch_out_fr, textvariable=batch_outdir_var,
             bg=C["field_bg"], fg=C["field_fg"], insertbackground=C["accent"],
             relief="flat", bd=0, font=("Segoe UI",9), highlightthickness=1,
             highlightbackground=C["field_border"], highlightcolor=C["accent"]
             ).grid(row=0, column=1, sticky="ew", padx=(0,4))
    def browse_batch_outdir():
        d = filedialog.askdirectory(title="Select output folder")
        if d: batch_outdir_var.set(d)
    RoundedButton(batch_out_fr, text="…", command=browse_batch_outdir, radius=10, padx=10).grid(row=0, column=2)
    tk.Label(batch_out_fr, text="(leave blank = same folder as source)",
             bg=C["card"], fg=C["muted"], font=("Segoe UI",7)
             ).grid(row=1, column=0, columnspan=3, sticky="w")

    # Batch log
    batch_log_var = tk.StringVar(value="")
    batch_status_lbl = tk.Label(batch_section, textvariable=batch_log_var,
                                bg=C["card"], fg=C["muted"],
                                font=("Segoe UI",8), wraplength=280, justify="left",
                                anchor="w")
    batch_status_lbl.grid(row=5, column=0, sticky="ew", pady=(2,2))

    batch_progress = ttk.Progressbar(batch_section, maximum=100)
    batch_progress.grid(row=6, column=0, sticky="ew", pady=(0,4))

    batch_run_btn = RoundedButton(batch_section, text="▶  Run Batch", fill=C["green"], hover="#55ee55", fg="#071307", radius=14, pady=6)
    _batch_card.grid_remove()  # hidden until toggled

    def do_batch():
        if not batch_queue:
            messagebox.showwarning("Batch", "Queue is empty — add files first.")
            return
        hdr = {k: v.get().strip() for k, v in hdr_vars.items()}
        outdir = batch_outdir_var.get().strip()
        jobs = []
        for item in batch_queue:
            src = item["src"]
            folder = outdir if outdir else os.path.dirname(src)
            stem = output_stem_for_item(item, hdr)
            dst = unique_output_path(folder, stem)
            jobs.append({"src": src, "dst": dst, "hdr": clean_unique_hdr(item.get("hdr", {}))})

        batch_run_btn.configure(state="disabled")
        batch_log_var.set("Starting batch…")
        batch_progress["value"] = 0

        log_lines = []
        def log_cb(msg):
            log_lines.append(msg)
            # Keep last 3 lines visible
            batch_log_var.set("\n".join(log_lines[-3:]))
            root.update_idletasks()

        def pct_cb(job_idx, n_jobs, file_pct, msg):
            overall = int((job_idx / n_jobs) * 100 + (file_pct / n_jobs))
            batch_progress["value"] = overall
            root.update_idletasks()

        def run():
            try:
                batch_convert(jobs, hdr_defaults=hdr,
                              progress_cb=pct_cb, log_cb=log_cb,
                              use_preview_sketches_for_excel=preview_excel_var.get(),
                              fit_sketches_to_cell=fit_sketches_to_cell_var.get())
                batch_progress["value"] = 100
                n_ok = sum(1 for j in jobs if os.path.isfile(j.get("dst", "")))
                root.after(0, lambda: messagebox.showinfo(
                    "Batch Complete", f"Batch finished.\n{n_ok} of {len(jobs)} file(s) processed."))
            except Exception as e:
                root.after(0, lambda: messagebox.showerror("Batch Error", str(e)))
            finally:
                root.after(0, lambda: batch_run_btn.configure(state="normal"))

        threading.Thread(target=run, daemon=True).start()

    batch_run_btn.configure(command=do_batch)
    batch_run_btn.grid(row=7, column=0, sticky="ew", pady=(0,4))

    # ══════════════════════════════════════════════════════════════════════
    # RIGHT PANEL
    # ══════════════════════════════════════════════════════════════════════
    right = tk.Frame(root, bg=C["bg"], padx=12, pady=12)
    right.grid(row=0, column=1, sticky="nsew")
    right.columnconfigure(0, weight=1)
    right.rowconfigure(2, weight=1)

    # ── stats bar ──
    _stats_card = RoundedCard(right, fill=C["card"], outline=C["stroke_soft"], radius=18, pad=8)
    _stats_card.grid(row=0, column=0, sticky="ew", pady=(0,8))
    stats_fr = _stats_card.inner
    stat_defs = [("Bars","bars"),("Diameters","dias"),("Tonnage","tons"),
                 ("Bar Type","type"),("Total Length","length"),
                 ("Qty (total)","qty"),("Weight (bar)","weight")]
    stat_vals = {}
    for i,(lbl,key) in enumerate(stat_defs):
        stats_fr.columnconfigure(i, weight=1)
        tk.Label(stats_fr, text=lbl, bg=C["card"], fg=C["muted"],
                 font=("Segoe UI",7), anchor="center").grid(row=0,column=i,sticky="ew",padx=4)
        lbl2 = tk.Label(stats_fr, text="—", bg=C["card"], fg=C["accent"],
                        font=("Segoe UI",10,"bold"), anchor="center")
        lbl2.grid(row=1,column=i,sticky="ew",padx=4)
        stat_vals[key] = lbl2

    # ── thumbnail strip ──
    _thumb_card = RoundedCard(right, fill=C["card"], outline=C["stroke_soft"], radius=18, pad=6)
    _thumb_card.grid(row=1, column=0, sticky="ew", pady=(0,8))
    thumb_fr = _thumb_card.inner
    THUMB_H = 56; THUMB_W = 80
    # Extra vertical room keeps the bar-mark labels visible above the scrollbar.
    thumb_canvas = tk.Canvas(thumb_fr, height=THUMB_H+32, bg=C["card"],
                             highlightthickness=0)
    thumb_canvas.pack(side="top", fill="x", expand=True)

    # Minimal horizontal scrollbar — 5 px tall, rounded thumb, no arrows
    _TH_H = 5
    _th_sb = tk.Canvas(thumb_fr, height=_TH_H, bg=C["card"],
                       highlightthickness=0, bd=0)
    _th_sb.pack(side="bottom", fill="x", pady=(4,0))

    def _th_draw(lo=None, hi=None):
        _th_sb.delete("all")
        try: lo, hi = (float(lo), float(hi)) if lo is not None else thumb_canvas.xview()
        except: return
        if hi - lo >= 1.0: return
        w = _th_sb.winfo_width() or 1
        tx = int(lo * w); tw = max(20, int((hi - lo) * w))
        r = _TH_H // 2; col = C["sb_thumb"]
        _th_sb.create_oval(tx,0,tx+_TH_H,_TH_H,       fill=col, outline="")
        _th_sb.create_oval(tx+tw-_TH_H,0,tx+tw,_TH_H, fill=col, outline="")
        _th_sb.create_rectangle(tx+r,0,tx+tw-r,_TH_H,  fill=col, outline="")

    _th_drag = {}
    def _th_press(e): _th_drag["x"]=e.x; _th_drag["lo"]=thumb_canvas.xview()[0]
    def _th_move(e):
        w=_th_sb.winfo_width() or 1; d=(e.x-_th_drag["x"])/w
        thumb_canvas.xview_moveto(_th_drag["lo"]+d); _th_draw()
    _th_sb.bind("<ButtonPress-1>", _th_press)
    _th_sb.bind("<B1-Motion>",     _th_move)
    _th_sb.bind("<Configure>",     lambda e: _th_draw())
    thumb_canvas.configure(xscrollcommand=_th_draw)
    thumb_inner = tk.Frame(thumb_canvas, bg=C["card"])
    thumb_window = thumb_canvas.create_window((0,0), window=thumb_inner, anchor="nw")
    def on_thumb_configure(e):
        thumb_canvas.configure(scrollregion=thumb_canvas.bbox("all"))
    thumb_inner.bind("<Configure>", on_thumb_configure)

    # ── main sketch preview ──
    _prev_card = RoundedCard(right, fill=C["card2"], outline=C["stroke"], radius=22, pad=8)
    _prev_card.grid(row=2, column=0, sticky="nsew")
    _prev_card.inner.columnconfigure(0, weight=1)
    # Keep only the sketch pane row expandable. If row 0 also has weight,
    # the title row can take the preview height and the sketch becomes hidden.
    _prev_card.inner.rowconfigure(0, weight=0)
    _prev_card.inner.rowconfigure(1, weight=1)
    _prev_card.inner.rowconfigure(2, weight=0)
    prev_fr = _prev_card.inner

    # Card header label
    tk.Label(prev_fr, text="Sketch Preview", bg=C["card2"], fg=C["accent"],
             font=("Segoe UI",8,"bold")).grid(row=0, column=0, sticky="w", padx=(2,0), pady=(0,5))

    pane = tk.PanedWindow(prev_fr, orient="horizontal", bg=C["card2"],
                          sashwidth=6, sashrelief="flat", bd=0, opaqueresize=True)
    pane.grid(row=1, column=0, sticky="nsew", pady=(2, 2))

    canvas = tk.Canvas(pane, bg=C["card3"], highlightthickness=0,
                       bd=0, width=560, height=340)
    pane.add(canvas, stretch="always")
    canvas.create_text(280, 170, text="Sketch preview will appear here",
                       fill=C["muted"], font=("Segoe UI", 11, "bold"))

    # Info panel on the right of the sketch
    _info_card = RoundedCard(pane, fill=C["card"], outline=C["stroke_soft"], radius=18, pad=8)
    pane.add(_info_card, stretch="never", width=172)
    info_fr = _info_card.inner
    info_fr.columnconfigure(0, weight=1)
    info_rows = [("Bar Mark","idx"),("Type","bar_type"),("Diameter","diameter"),
                 ("No. Elements","nb_elem"),("No. Steel","nb_steel"),
                 ("Total Qty","nb_total"),("Total Length","total_len"),
                 ("Unit Weight","unit_wt"),("Total Weight","total_wt"),
                 ("Mandrin","mandrin")]
    info_labels = {}
    for i,(lbl,key) in enumerate(info_rows):
        tk.Label(info_fr, text=lbl, bg=C["card"], fg=C["muted"],
                 font=("Segoe UI",7), anchor="w").grid(row=i*2, column=0, sticky="w", pady=(4,0))
        vl = tk.Label(info_fr, text="—", bg=C["card"], fg=C["accent"],
                      font=("Segoe UI",9,"bold"), anchor="w")
        vl.grid(row=i*2+1, column=0, sticky="ew")
        info_labels[key] = vl

    # Nav bar
    nav_fr = tk.Frame(prev_fr, bg=C["card2"])
    nav_fr.grid(row=2, column=0, sticky="ew", pady=(8,0))
    nav_fr.columnconfigure(1, weight=1)
    RoundedButton(nav_fr, text="◀ Prev", command=lambda: render_preview(ps["idx"]-1), radius=12, padx=12, pady=6).grid(row=0,column=0,padx=(0,8))
    nav_label = tk.Label(nav_fr, text="No file loaded", bg=C["card2"],
                         fg=C["muted"], font=("Segoe UI",8), anchor="center")
    nav_label.grid(row=0, column=1, sticky="ew")
    RoundedButton(nav_fr, text="Next ▶", command=lambda: render_preview(ps["idx"]+1), radius=12, padx=12, pady=6).grid(row=0,column=2,padx=(8,0))

    root.bind("<Left>",  lambda e: render_preview(ps["idx"]-1))
    root.bind("<Right>", lambda e: render_preview(ps["idx"]+1))

    # ══════════════════════════════════════════════════════════════════════
    # PREVIEW STATE & RENDERING
    # ══════════════════════════════════════════════════════════════════════
    ps = {"bars":[],"idx":0,"wmf_map":{},"photo":None,"thumb_photos":[]}

    def render_preview(idx):
        if not ps["bars"]: return
        idx = max(0, min(idx, len(ps["bars"])-1))
        ps["idx"] = idx
        bar  = ps["bars"][idx]
        dims = bar.get("dims", {})
        dia  = bar.get("diameter", 0)
        dc   = dia_color(dia)
        qty  = bar.get("nb_total", 0)
        tlen = bar.get("total_len", 0)
        uwt  = bar_weight(dia, tlen, 1)
        twt  = bar_weight(dia, tlen, qty)
        has_wmf = bar.get("src_row",-1) in ps["wmf_map"]

        nav_label.configure(
            text=f"Bar {bar.get('idx','')}  –  {bar.get('bar_type','')}  │  "
                 f"{'[WMF ✓]' if has_wmf else '[generated]'}  │  {idx+1} / {len(ps['bars'])}")

        # Update info panel
        info_labels["idx"].configure(      text=str(bar.get("idx","")),       fg=dc)
        info_labels["bar_type"].configure( text=str(bar.get("bar_type","")),  fg=dc)
        info_labels["diameter"].configure( text=f"T{dia}" if dia else "—",    fg=dc)
        info_labels["nb_elem"].configure(  text=str(bar.get("nb_elem","")))
        info_labels["nb_steel"].configure( text=str(bar.get("nb_steel","")))
        info_labels["nb_total"].configure( text=str(qty))
        info_labels["total_len"].configure(text=f"{tlen} mm" if tlen else "—")
        info_labels["unit_wt"].configure(  text=f"{uwt:.3f} kg" if uwt else "—")
        info_labels["total_wt"].configure( text=f"{twt:.2f} kg" if twt else "—")
        info_labels["mandrin"].configure(  text=str(bar.get("mandrin","")) or "—")

        # Update current stats bar
        stat_vals["type"].configure(  text=bar.get("bar_type","—"))
        stat_vals["length"].configure(text=f"{tlen} mm" if tlen else "—")
        stat_vals["qty"].configure(   text=str(qty))
        stat_vals["weight"].configure(text=f"{twt:.2f} kg" if twt else "—")

        # Scroll thumbnail strip to current
        if ps["thumb_photos"]:
            total_w = len(ps["bars"]) * (THUMB_W+4)
            frac = idx / max(len(ps["bars"])-1, 1)
            thumb_canvas.xview_moveto(max(0, frac - 0.3))

        # Draw sketch
        cw = max(canvas.winfo_width(),  400)
        ch = max(canvas.winfo_height(), 200)
        canvas.delete("all")
        canvas.configure(bg=C["card3"])

        try:
            from PIL import Image, ImageTk, ImageFilter
            buf = draw_sketch(dims, total_len=tlen)
            if buf:
                raw = Image.open(buf).convert("RGBA")
                raw = _trim_transparent_pil_image(raw, pad_px=12)
                # Tint lines with diameter color
                r,g,b = int(dc[1:3],16), int(dc[3:5],16), int(dc[5:7],16)
                data = raw.load()
                for py in range(raw.height):
                    for px in range(raw.width):
                        pr,pg,pb,pa = data[px,py]
                        if pa > 10 and pr > 100:  # red pixels -> recolor
                            data[px,py] = (r,g,b,pa)
                raw.thumbnail((cw-32, ch-32), Image.LANCZOS)
                bg = Image.new("RGBA",(cw,ch),(
                    int(C["card3"][1:3],16),
                    int(C["card3"][3:5],16),
                    int(C["card3"][5:7],16),255))
                bg.paste(raw,((cw-raw.width)//2,(ch-raw.height)//2), raw)
                photo = ImageTk.PhotoImage(bg)
                ps["photo"] = photo
                canvas.create_image(0, 0, anchor="nw", image=photo)
                if has_wmf:
                    canvas.create_rectangle(cw-96,4,cw-4,20,fill=C["card"],outline=C["green"])
                    canvas.create_text(cw-8,12,anchor="e",
                                       text="⚡ WMF source",fill=C["green"],font=("Segoe UI",7))
            else:
                canvas.create_text(cw//2,ch//2,
                    text="Straight bar  —  no sketch",
                    fill=C["muted"], font=("Segoe UI",11))
        except Exception as ex:
            canvas.create_text(cw//2,ch//2,text=f"Preview unavailable\n{ex}",
                               fill=C["muted"],font=("Segoe UI",9))

    def on_resize(e):
        if ps["bars"]: root.after(60, lambda: render_preview(ps["idx"]))
    canvas.bind("<Configure>", on_resize)

    def build_thumbnails():
        for w in thumb_inner.winfo_children(): w.destroy()
        ps["thumb_photos"] = []
        try:
            from PIL import Image, ImageTk, ImageDraw
        except: return
        for i, bar in enumerate(ps["bars"]):
            dims = bar.get("dims",{})
            dia  = bar.get("diameter",0)
            dc   = dia_color(dia)
            r,g,b= int(dc[1:3],16),int(dc[3:5],16),int(dc[5:7],16)
            bg_hex = C["card2"]
            bg_rgb = (int(bg_hex[1:3],16),int(bg_hex[3:5],16),int(bg_hex[5:7],16))

            bg_img = Image.new("RGBA",(THUMB_W,THUMB_H),bg_rgb+(255,))
            buf = draw_sketch(dims, total_len=bar.get("total_len", 0))
            if buf:
                raw = Image.open(buf).convert("RGBA")
                raw = _trim_transparent_pil_image(raw, pad_px=8)
                data = raw.load()
                for py in range(raw.height):
                    for px in range(raw.width):
                        pr,pg,pb,pa = data[px,py]
                        if pa>10 and pr>100: data[px,py]=(r,g,b,pa)
                raw.thumbnail((THUMB_W-8, THUMB_H-8), Image.LANCZOS)
                bg_img.paste(raw,((THUMB_W-raw.width)//2,(THUMB_H-raw.height)//2),raw)
            # Color indicator strip at bottom
            draw = ImageDraw.Draw(bg_img)
            draw.rectangle([0,THUMB_H-4,THUMB_W,THUMB_H], fill=(r,g,b,255))

            photo = ImageTk.PhotoImage(bg_img)
            ps["thumb_photos"].append(photo)

            idx_copy = i
            fr = tk.Frame(thumb_inner, bg=C["card2"], padx=1, pady=1, cursor="hand2")
            fr.pack(side="left", padx=2, pady=4)
            lbl = tk.Label(fr, image=photo, bg=C["card2"], relief="flat", bd=0)
            lbl.pack()
            tk.Label(fr, text=str(bar.get("idx",i+1)), bg=C["card2"],
                     fg=dc, font=("Segoe UI",7)).pack()
            def on_click(e, ix=idx_copy):
                render_preview(ix)
            lbl.bind("<Button-1>", on_click)
            fr.bind("<Button-1>",  on_click)

    def load_preview(path):
        try:
            bars_df = parse_nomenclature(path)
            sketch_map = extract_source_sketches(path)
            bars = bars_df.to_dict("records")
            ps["bars"] = bars; ps["wmf_map"] = sketch_map; ps["idx"] = 0

            dias  = sorted(set(int(b.get("diameter",0)) for b in bars if b.get("diameter",0)))
            total_tons = sum(bar_weight(b.get("diameter",0),
                                        b.get("total_len",0),
                                        b.get("nb_total",0))
                             for b in bars) / 1000

            def upd():
                stat_vals["bars"].configure( text=str(len(bars)))
                stat_vals["dias"].configure( text="  ".join(f"T{d}" for d in dias))
                stat_vals["tons"].configure( text=f"{total_tons:.3f} t")
                build_thumbnails()
                render_preview(0)
            root.after(0, upd)
        except Exception as ex:
            root.after(0, lambda: nav_label.configure(text=f"Error: {ex}"))

    # ══════════════════════════════════════════════════════════════════════
    # RESTORE + ACTIONS
    # ══════════════════════════════════════════════════════════════════════
    if cfg.get("src"):  src_var.set(cfg["src"])
    if cfg.get("dst"):  dst_var.set(cfg["dst"])
    # Restore only shared/non-unique header fields from config.
    # Unique fields must not be retained after closing/reopening the app/exe.
    for key, v in hdr_vars.items():
        if key in HEADER_UNIQUE_KEYS:
            v.set("")
        elif cfg.get(key):
            v.set(cfg[key])
    if cfg.get("src") and os.path.isfile(cfg["src"]):
        drop_lbl.configure(text=f"✔  {os.path.basename(cfg['src'])}",
                           bg=C["card"], fg=C["green"])
        root.after(300, lambda: threading.Thread(
            target=lambda: load_preview(cfg["src"]), daemon=True).start())

    def _refresh_sidebar_bg():
        _sb_shell.configure(bg=C["sidebar"])
        _sb_canvas.configure(bg=C["sidebar"])
        _sb_bar.configure(bg=C["sidebar"])
        sidebar.configure(bg=C["sidebar"])
        _sb_draw_thumb()
        right.configure(bg=C["bg"])

    def persist():
        # Persist only application settings and shared/non-unique header fields.
        # Do not save bbs_name, bbs_no, sd_no, bbs_rev, or sd_rev.
        config_data = {
            "src": src_var.get(),
            "dst": dst_var.get(),
            "recent": recent_files,
            "presets": presets,
            "excel_preview_sketches": preview_excel_var.get(),
            "fit_sketches_to_cell": fit_sketches_to_cell_var.get(),
        }
        config_data.update({
            k: v.get() for k, v in hdr_vars.items()
            if k not in HEADER_UNIQUE_KEYS
        })
        save_config(config_data)

    def do_convert():
        src=src_var.get().strip(); dst=dst_var.get().strip()
        if not src: return messagebox.showwarning("Missing input","Please select a source file.")
        if not dst: return messagebox.showwarning("Missing output","Please specify an output path.")
        if not os.path.isfile(src): return messagebox.showerror("Not found",f"Cannot find:\n{src}")
        hdr={k:v.get().strip() for k,v in hdr_vars.items()}
        persist(); convert_btn.configure(state="disabled")
        def run():
            try:
                convert(src,dst,hdr,
                    progress_cb=lambda p,m:(progress.__setitem__("value",p),
                                            status_var.set(m),root.update_idletasks()),
                    use_preview_sketches_for_excel=preview_excel_var.get(),
                    fit_sketches_to_cell=fit_sketches_to_cell_var.get())
                messagebox.showinfo("Success",f"Saved to:\n{dst}")
            except Exception as e:
                messagebox.showerror("Error",str(e)); status_var.set(f"Error: {e}")
            finally:
                convert_btn.configure(state="normal")
        threading.Thread(target=run,daemon=True).start()
    convert_btn.configure(command=do_convert)
    # Propagate mousewheel to sidebar scroll from any child widget
    def _bind_mousewheel_recursive(widget):
        widget.bind("<MouseWheel>", _sb_mousewheel, add="+")
        widget.bind("<Button-4>",   _sb_mousewheel, add="+")
        widget.bind("<Button-5>",   _sb_mousewheel, add="+")
        for child in widget.winfo_children():
            _bind_mousewheel_recursive(child)
    root.after(200, lambda: _bind_mousewheel_recursive(sidebar))

    root.protocol("WM_DELETE_WINDOW", lambda:(persist(),root.destroy()))

    root.update_idletasks()
    sw,sh = root.winfo_screenwidth(), root.winfo_screenheight()
    w = cfg.get("win_w", 1200); h = cfg.get("win_h", 700)
    root.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
    root.mainloop()

# ── entry point ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="BBS Transformer — Armabeton XLS → English Bar Bending Schedule XLSX",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  Single file:
    python bbs_transformer.py src.xls out.xlsx
    python bbs_transformer.py src.xls out.xlsx --fit-sketches-to-cell
    python bbs_transformer.py src.xls out.xlsx --preview-sketches-excel --fit-sketches-to-cell

  Batch (explicit pairs):
    python bbs_transformer.py --batch src1.xls src2.xls --outdir ./output

  Batch (folder scan):
    python bbs_transformer.py --folder ./nomenclatures --outdir ./output
""",
    )
    parser.add_argument("files", nargs="*",
                        help="Source file(s) — positional. "
                             "Two args = single-file mode (src dst). "
                             "One or more with --batch = batch mode.")
    parser.add_argument("--batch", action="store_true",
                        help="Treat all positional args as source files (batch mode).")
    parser.add_argument("--folder", metavar="DIR",
                        help="Scan a folder for .xls/.xlsx files and batch-convert all.")
    parser.add_argument("--outdir", metavar="DIR",
                        help="Output directory for batch mode (default: same as source).")
    parser.add_argument("--preview-sketches-excel", action="store_true",
                        help="Write generated preview sketches into Excel instead of source WMF sketches; fitted/centered in SKETCH cells.")
    parser.add_argument("--source-sketches-excel", action="store_true",
                        help="Use original source WMF sketches when available (default).")
    parser.add_argument("--fit-sketches-to-cell", "--exact-fit-sketches", action="store_true",
                        help="Stretch sketches to exactly fill each SKETCH cell instead of centered aspect-ratio fit.")
    # Header fields
    parser.add_argument("--project",    default="", metavar="TEXT")
    parser.add_argument("--bbs-name",   default="", metavar="TEXT")
    parser.add_argument("--bbs-no",     default="", metavar="TEXT")
    parser.add_argument("--sd-no",      default="", metavar="TEXT")
    parser.add_argument("--office",     default="", metavar="TEXT")
    parser.add_argument("--bbs-rev",    default="00", metavar="REV")
    parser.add_argument("--sd-rev",     default="00", metavar="REV")
    parser.add_argument("--prepared-by",default="",  metavar="INITIALS")
    parser.add_argument("--checked-by", default="",  metavar="INITIALS")

    args = parser.parse_args()

    hdr = {
        "project_name": args.project,
        "bbs_name":     args.bbs_name,
        "bbs_no":       args.bbs_no,
        "sd_no":        args.sd_no,
        "office":       args.office,
        "bbs_rev":      args.bbs_rev,
        "sd_rev":       args.sd_rev,
        "prepared_by":  args.prepared_by,
        "checked_by":   args.checked_by,
    }

    # ── single-file shortcut (legacy: exactly 2 positional args, no --batch/--folder) ──
    if not args.batch and not args.folder and len(args.files) == 2:
        src, dst = args.files
        print(f"Converting {src} → {dst}")
        convert(src, dst, hdr,
                progress_cb=lambda p, m: print(f"  [{p:3d}%] {m}"),
                use_preview_sketches_for_excel=args.preview_sketches_excel,
                fit_sketches_to_cell=args.fit_sketches_to_cell)
        sys.exit(0)

    # ── batch mode ──
    srcs = list(args.files)

    if args.folder:
        folder = args.folder
        if not os.path.isdir(folder):
            print(f"Error: --folder path not found: {folder}", file=sys.stderr)
            sys.exit(1)
        found = [os.path.join(folder, f) for f in os.listdir(folder)
                 if f.lower().endswith((".xls", ".xlsx")) and not f.startswith("~$")]
        srcs.extend(found)
        print(f"Found {len(found)} file(s) in {folder}")

    if not srcs:
        # No files and no folder → launch GUI
        launch_gui()
        sys.exit(0)

    outdir = args.outdir
    if outdir and not os.path.isdir(outdir):
        os.makedirs(outdir, exist_ok=True)
        print(f"Created output folder: {outdir}")

    jobs = []
    for src in srcs:
        if outdir:
            dst = os.path.join(outdir,
                               os.path.splitext(os.path.basename(src))[0] + "_BBS.xlsx")
        else:
            dst = os.path.splitext(src)[0] + "_BBS.xlsx"
        jobs.append({"src": src, "dst": dst, "hdr": {}})

    print(f"\nBatch: {len(jobs)} file(s) to convert")
    print("-" * 50)

    def cli_pct(job_idx, n_jobs, file_pct, msg):
        print(f"  [{file_pct:3d}%] {msg}")

    results = batch_convert(jobs, hdr_defaults=hdr,
                            progress_cb=cli_pct,
                            log_cb=lambda m: print(m),
                            use_preview_sketches_for_excel=args.preview_sketches_excel,
                            fit_sketches_to_cell=args.fit_sketches_to_cell)

    n_ok  = sum(r["ok"] for r in results)
    n_err = len(results) - n_ok
    sys.exit(0 if n_err == 0 else 1)
